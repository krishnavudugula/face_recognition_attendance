from flask import Flask, request, jsonify, send_from_directory, send_file, make_response
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, time, timedelta
from threading import Thread
import numpy as np
import cv2
import base64
import os
import io
import pandas as pd
import pickle
import bcrypt
import pytz
import math
import json
import requests
from functools import wraps

# 🔴 FCM SERVER-SIDE: Modern Firebase Admin SDK for device wakeups
try:
    import firebase_admin
    from firebase_admin import credentials, messaging
    FIREBASE_ADMIN_AVAILABLE = True
    
    # Initialize Firebase Admin SDK if not already initialized
    if not firebase_admin._apps:
        # Look for service account JSON in multiple locations
        service_account_paths = [
            'serviceAccountKey.json',
            os.path.join(os.path.dirname(__file__), 'serviceAccountKey.json'),
            os.getenv('GOOGLE_APPLICATION_CREDENTIALS'),
        ]
        
        firebase_initialized = False
        for path in service_account_paths:
            if path and os.path.exists(path):
                try:
                    cred = credentials.Certificate(path)
                    firebase_admin.initialize_app(cred)
                    print(f"✅ Firebase Admin SDK initialized with: {path}")
                    firebase_initialized = True
                    break
                except Exception as e:
                    print(f"⚠️ Failed to initialize Firebase with {path}: {e}")
        
        if not firebase_initialized:
            try:
                # Try Application Default Credentials (ADC) from environment
                firebase_admin.initialize_app()
                print("✅ Firebase Admin SDK initialized with Application Default Credentials")
                firebase_initialized = True
            except Exception as e:
                print(f"⚠️ Firebase Admin SDK not initialized: {e}")
                print("   Location tracking FCM pings will NOT work.")
                print("   Either:")
                print("   1. Download service account JSON from Firebase Console")
                print("   2. Save it as 'serviceAccountKey.json' in project root")
                print("   3. OR set GOOGLE_APPLICATION_CREDENTIALS environment variable")
    else:
        print("✅ Firebase Admin SDK already initialized")
        
except ImportError:
    FIREBASE_ADMIN_AVAILABLE = False
    print("⚠️ firebase_admin not installed. Install with: pip install firebase-admin")

# 🔴 BACKGROUND SCHEDULER: Ping devices every 10 minutes for fresh location
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.interval import IntervalTrigger
    APSCHEDULER_AVAILABLE = True
except ImportError:
    APSCHEDULER_AVAILABLE = False
    print("⚠️ apscheduler not installed. Install with: pip install apscheduler")
    print("   (App will work but background device pings will be disabled)")
# Target Location: Position: 17.937351,79.849383 Radius: 242.07 Meters(Testing)
TARGET_LAT = 17.937351
TARGET_LON = 79.849383
ALLOWED_RADIUS_KM = 0.24207  # 242.07 meters
LOCATION_ENFORCEMENT_ENABLED = True  # Set to False to allow attendance from anywhere (testing mode)
GPS_ACCURACY_BUFFER_M = 50  # 🔴 CRITICAL: Account for GPS error (~5-20m device error + buffer for safety)
# Effective boundary = ALLOWED_RADIUS_KM - GPS_ACCURACY_BUFFER_M
# This prevents false IN BOUNDS when physically outside but GPS reports as inside

def haversine(lat1, lon1, lat2, lon2):
    # Radius of the Earth in km
    R = 6371.0

    # Convert latitude and longitude from degrees to radians
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)

    # Differences in coordinates
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad

    # Haversine formula
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    distance = R * c
    return distance

import urllib.request

# ⚠️ DEPRECATED — This class uses a dead FCM API that was shut down June 2024
# DO NOT USE — Use ping_device_via_fcm() instead (Firebase Admin SDK)
#
# class FirebaseMessagingService:
#     """DEAD — DO NOT USE — Legacy FCM API is no longer available"""
#     # Kept for reference only. The https://fcm.googleapis.com/fcm/send endpoint was shut down.

# ════════════════════════════════════════════════════════════════════
# 🔴 FCM SERVER-SIDE: Ping devices every 10 minutes for fresh location
# ════════════════════════════════════════════════════════════════════

def ping_device_via_fcm(fcm_token, action="heartbeat"):
    """
    🔴 CRITICAL: Ping device via FCM to trigger location update.
    
    This is what makes FCM actually work. Without this,
    devices just wait forever for a server signal.
    
    Firebase Admin SDK method (modern, reliable):
    - Wakes device even in Doze mode (priority="high")
    - No HTTP auth key needed (uses service account)
    - Sends data-only message (no notification popup)
    
    Args:
        fcm_token: Device's FCM registration token
        action: "heartbeat" | "sync" | "urgent"
    
    Returns:
        bool: True if message sent successfully
    """
    if not FIREBASE_ADMIN_AVAILABLE:
        print("⚠️ Firebase Admin SDK not available, skipping FCM ping")
        print("   Devices will not receive wakeup signals!")
        return False
    
    if not fcm_token:
        print("⚠️ No FCM token provided")
        return False
    
    try:
        # 🔴 Data-only message (no notification popup)
        # This wakes the device silently and triggers onMessageReceived()
        message = messaging.Message(
            data={
                "action": action,
                "timestamp": datetime.utcnow().isoformat()
            },
            token=fcm_token,
            android=messaging.AndroidConfig(
                priority="high",  # ← MUST be "high" to wake Doze mode!
                ttl=300  # 5 minute timeout
            )
        )
        
        response = messaging.send(message)
        print(f"✅ FCM ping sent: {fcm_token[:20]}... (action: {action}, msg_id: {response[:20]}...)")
        return True
        
    except Exception as e:
        print(f"❌ FCM ping FAILED: {str(e)}")
        print(f"   Token: {fcm_token[:30]}...")
        import traceback
        traceback.print_exc()
        return False


def ping_all_active_devices():
    """
    🔴 BACKGROUND TASK: Ping all logged-in devices every 10 minutes.
    
    Called by APScheduler background job every 10 minutes.
    Ensures devices stay alive and report fresh location.
    
    Flow:
    1. Find all users with active LivePresence (logged in, last seen < 15 min)
    2. Get their FCM tokens
    3. Send "heartbeat" ping via FCM
    4. Device wakes → sends fresh location
    5. Device sleeps until next ping
    
    Result: 30x more battery efficient than 10-second device heartbeat
    """
    try:
        print("\n🔔 [FCM BATCH PING] Starting device ping batch...")
        now_utc = datetime.utcnow()
        cutoff = now_utc - timedelta(minutes=15)  # Active if seen in last 15 min
        
        active_presences = LivePresence.query.filter(
            LivePresence.last_seen > cutoff
        ).all()
        
        if not active_presences:
            print("💤 No active users to ping (no LivePresence records in last 15 min)")
            return
        
        print(f"📍 Found {len(active_presences)} active user(s)")
        total_pinged = 0
        success_count = 0
        
        for presence in active_presences:
            # Get all active FCM tokens for this user
            tokens = FCMToken.query.filter_by(
                user_id=presence.user_id,
                is_active=True
            ).all()
            
            print(f"   User: {presence.user_id}, Tokens: {len(tokens)}")
            
            for token_obj in tokens:
                total_pinged += 1
                if ping_device_via_fcm(token_obj.fcm_token, "heartbeat"):
                    success_count += 1
        
        print(f"\n✅ [FCM BATCH COMPLETE] {success_count}/{total_pinged} devices pinged successfully\n")
        
    except Exception as e:
        print(f"\n❌ [FCM BATCH ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        print()

def ping_inactive_devices():
    """
    🔴 AGGRESSIVE PING: Wake devices that haven't reported in 10 minutes.
    
    Runs every 5 minutes. More targeted than ping_all_active_devices().
    If a device hasn't sent a location update in 10 minutes, it's likely
    sleeping or stuck. Send it an FCM wakeup ping.
    """
    try:
        print("\n🔔 [INACTIVE DEVICE PING] Checking for devices without recent updates...")
        
        # Get devices that haven't updated in the last 10 minutes
        cutoff = datetime.utcnow() - timedelta(minutes=10)
        inactive_devices = db.session.execute(
            db.text("""
                SELECT DISTINCT user_id FROM location_log 
                WHERE timestamp > :cutoff 
                GROUP BY user_id 
                HAVING MAX(timestamp) < :cutoff2
            """),
            {"cutoff": datetime.utcnow() - timedelta(minutes=15), "cutoff2": cutoff}
        ).fetchall()
        
        if not inactive_devices:
            print("✅ No inactive devices to wake up")
            return
        
        print(f"📍 Found {len(inactive_devices)} devices with no recent updates")
        success_count = 0
        
        for (user_id,) in inactive_devices:
            # Get active FCM tokens for this user
            tokens = FCMToken.query.filter_by(
                user_id=user_id,
                is_active=True
            ).all()
            
            for token_obj in tokens:
                if ping_device_via_fcm(token_obj.fcm_token, "wakeup"):
                    success_count += 1
        
        print(f"✅ [INACTIVE PING COMPLETE] Woke up {success_count} devices\n")
        
    except Exception as e:
        print(f"\n❌ [INACTIVE PING ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        print()

def auto_mark_incomplete_attendance():
    """
    🔴 END-OF-DAY AUTO-MARKING AT 6:00 PM: Mark faculty ABSENT if they didn't mark twice.
    
    Runs daily at 6:00 PM IST (18:00).
    Finds all AttendanceLogs with morning check-in but NO evening check-out by 6 PM.
    
    Logic:
    - If morning mark exists but NO evening mark by 6 PM → Status = "Absent"
    - This ensures faculty are marked ABSENT if they only mark once
    - No "HD" (half day) status - straight ABSENT
    """
    try:
        now_utc = datetime.utcnow()
        local_tz = pytz.timezone('Asia/Kolkata')
        now_local = now_utc.astimezone(local_tz)
        today_str = now_local.strftime('%Y-%m-%d')
        
        # Find all incomplete attendances for today (checked in but not checked out)
        incomplete_logs = AttendanceLog.query.filter_by(
            date=today_str
        ).filter(
            AttendanceLog.time_out == None  # No check-out yet
        ).all()
        
        if not incomplete_logs:
            print("✅ No incomplete attendances to auto-mark")
            return
        
        auto_marked_count = 0
        for log in incomplete_logs:
            first_status = (log.check_in_status or log.status or "").strip()
            
            # If they marked in morning but didn't mark in evening → ABSENT
            if first_status in {"Present", "Late Permission"}:
                log.status = "Absent"
                log.check_out_status = "Absent"
                log.check_out_period = "18:00-24:00 (Auto-marked ABSENT - No evening mark)"
                log.time_out = now_local.strftime('%H:%M:%S')
                log.timestamp_out = now_utc
                print(f"  📝 Auto-marked {log.user_id} as ABSENT (no evening mark by 6 PM)")
                auto_marked_count += 1
            elif first_status in {"Absent", "Didn't Mark"}:
                # Already absent, no action needed
                pass
            elif first_status in {"HD", "EP", "Early Permission"}:
                # Already marked as half-day or early departure, no action needed
                pass
        
        if auto_marked_count > 0:
            db.session.commit()
            print(f"✅ Auto-marked {auto_marked_count} faculty as ABSENT (incomplete double marking)")
        else:
            print("💤 No attendances needed auto-marking")
            
    except Exception as e:
        print(f"⚠️ Auto-mark attendance error: {str(e)}")
        db.session.rollback()


# --- Custom Face Recognition & Liveness Logic ---
class FaceSystem:
    def __init__(self):
        """
        Face recognition using pure OpenCV (YuNet + SFace).
        - Extremely lightweight (~40MB total)
        - No deepface or tensorflow required
        - 99.60% accuracy (SFace model)
        - Perfect for PythonAnywhere free tier
        """
        self.models_loaded = False

        # Try multiple model locations for PythonAnywhere compatibility
        possible_dirs = [
            'data/models',
            '/tmp/models',
            os.path.expanduser('~/models'),
            os.path.join(os.getcwd(), 'models')
        ]

        models_dir = None
        for dir_path in possible_dirs:
            try:
                if not os.path.exists(dir_path):
                    os.makedirs(dir_path, exist_ok=True)
                    print(f"✓ Created models directory: {dir_path}")
                models_dir = dir_path
                break
            except Exception as e:
                print(f"✗ Cannot create {dir_path}: {e}")
                continue

        if not models_dir:
            print("CRITICAL: No writable directory found for models!")
            self.models_loaded = False
            return

        self.detector_path = os.path.join(models_dir, 'face_detection_yunet_2023mar.onnx')
        self.recognizer_path = os.path.join(models_dir, 'face_recognition_sface_2021dec.onnx')

        print(f"Using models directory: {models_dir}")
        print(f"Detector path: {self.detector_path}")
        print(f"Recognizer path: {self.recognizer_path}")

        # Auto-download models if they don't exist
        try:
            if not os.path.exists(self.detector_path):
                print("📥 Downloading YuNet face detection model...")
                try:
                    urllib.request.urlretrieve(
                        "https://github.com/opencv/opencv_zoo/raw/main/models/face_detection_yunet/face_detection_yunet_2023mar.onnx",
                        self.detector_path,
                        reporthook=lambda a,b,c: print(f"  Downloaded {a}/{c} bytes", end='\r')
                    )
                    print("✓ YuNet model downloaded successfully")
                except Exception as e:
                    print(f"✗ Failed to download YuNet: {e}")
                    raise

            if not os.path.exists(self.recognizer_path):
                print("📥 Downloading SFace face recognition model...")
                try:
                    urllib.request.urlretrieve(
                        "https://github.com/opencv/opencv_zoo/raw/main/models/face_recognition_sface/face_recognition_sface_2021dec.onnx",
                        self.recognizer_path,
                        reporthook=lambda a,b,c: print(f"  Downloaded {a}/{c} bytes", end='\r')
                    )
                    print("✓ SFace model downloaded successfully")
                except Exception as e:
                    print(f"✗ Failed to download SFace: {e}")
                    raise

            print("✓ Models verified")
            print(f"  Detector: {os.path.exists(self.detector_path)} ({os.path.getsize(self.detector_path) if os.path.exists(self.detector_path) else 0} bytes)")
            print(f"  Recognizer: {os.path.exists(self.recognizer_path)} ({os.path.getsize(self.recognizer_path) if os.path.exists(self.recognizer_path) else 0} bytes)")

            # Initialize OpenCV Face Detector (YuNet)
            print("🔧 Initializing YuNet detector...")
            self.detector = cv2.FaceDetectorYN.create(
                self.detector_path,
                "",
                (320, 320),
                0.5,  # Score threshold
                0.3,  # NMS threshold
                5000  # Top K
            )
            if self.detector is None:
                raise Exception("YuNet detector initialization returned None")
            print("✓ YuNet detector initialized")

            # Initialize OpenCV Face Recognizer (SFace)
            print("🔧 Initializing SFace recognizer...")
            self.recognizer = cv2.FaceRecognizerSF.create(self.recognizer_path, "")
            if self.recognizer is None:
                raise Exception("SFace recognizer initialization returned None")
            print("✓ SFace recognizer initialized")

            self.models_loaded = True
            print("✅ OpenCV YuNet & SFace Models Loaded Successfully!")

        except Exception as e:
            print(f"❌ CRITICAL ERROR loading OpenCV models: {e}")
            import traceback
            traceback.print_exc()
            self.models_loaded = False

    def check_image_quality(self, img):
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        avg_brightness = np.mean(gray)
        if avg_brightness < 40: return False, "Image too dark. Ensure good lighting."
        if avg_brightness > 220: return False, "Image too bright. Avoid direct glare."
        laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()
        if laplacian_var < 50: return False, "Image too blurry. Hold still."
        return True, "OK"

    def get_face_encoding(self, image_stream):
        if not self.models_loaded: return None, None

        if isinstance(image_stream, (bytes, bytearray)):
             file_bytes = np.frombuffer(image_stream, np.uint8)
        else:
             file_bytes = np.frombuffer(image_stream.read(), np.uint8)

        img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if img is None: return None, None

        passed, reason = self.check_image_quality(img)
        if not passed: print(f"Quality Check Failed: {reason}")

        try:
            height, width, _ = img.shape
            self.detector.setInputSize((width, height))

            # Detect faces
            _, faces = self.detector.detect(img)
            if faces is None:
                return None, None

            # Process largest face
            largest_face = max(faces, key=lambda f: f[2] * f[3])

            # Align and Extract Feature
            aligned_face = self.recognizer.alignCrop(img, largest_face)
            feat = self.recognizer.feature(aligned_face) # 1x128 array

            embedding = feat[0].astype(np.float64)

            location = {
                "top": int(largest_face[1]),
                "right": int(largest_face[0] + largest_face[2]),
                "bottom": int(largest_face[1] + largest_face[3]),
                "left": int(largest_face[0])
            }

            return embedding, location

        except Exception as e:
            print(f"Face encoding error: {e}")
            return None, None

    def compare_faces(self, known_encodings_dict, unknown_encoding, threshold=0.40):
        # OpenCV Cosine returns SIMILARITY (higher is better).
        if not known_encodings_dict: return None

        try:
            unknown_encoding = np.array(unknown_encoding, dtype=np.float32).reshape(1, 128)
        except Exception:
            return None

        best_match_id = None
        max_score = -float('inf')

        for user_id, encodings_list in known_encodings_dict.items():
            for known_enc in encodings_list:
                try:
                    if known_enc is None: continue
                    known = np.array(known_enc, dtype=np.float32).reshape(1, 128)

                    # Calculate Cosine SIMILARITY (Higher score = better match)
                    score = self.recognizer.match(known, unknown_encoding, cv2.FaceRecognizerSF_FR_COSINE)

                    # We want to find the HIGHEST similarity score
                    if score > max_score:
                        max_score = score
                        best_match_id = user_id
                except Exception:
                    continue

        # If the highest similarity is greater than our threshold, it's a match!
        if best_match_id and max_score >= threshold:
            print(f"  ✓ Face match: {best_match_id} (similarity: {max_score:.4f})")
            return best_match_id

        if best_match_id:
            print(f"  ✗ No match (closest: {best_match_id}, similarity: {max_score:.4f}, threshold required: {threshold})")
        return None

# Initialize face recognition system
face_system = FaceSystem()

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app,
     origins=['*'],  # Allow all origins (mobile apps and browsers)
     supports_credentials=True,
     expose_headers=[
         'Content-Disposition',
         'Content-Type',
         'Content-Length',
         'Access-Control-Allow-Origin',
         'Access-Control-Allow-Headers',
         'Access-Control-Allow-Methods'
     ],
     allow_headers=[
         'Content-Type',
         'Authorization',
         'Accept',
         'ngrok-skip-browser-warning'
     ],
     methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])

# 🔴 CRITICAL: Handle ngrok browser warning bypass
@app.before_request
def handle_ngrok_preflight():
    """Handle ngrok OPTIONS preflight requests gracefully"""
    if request.method == 'OPTIONS':
        print(f"[CORS] OPTIONS preflight from: {request.origin}")
        response = make_response()
        response.headers['Access-Control-Allow-Origin'] = request.origin if request.origin else '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, Accept, ngrok-skip-browser-warning'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.status_code = 200
        return response

# Database Configuration
# Use Environment Variable for DB URI if available, else fallback to SQLite
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///' + os.path.join(basedir, 'face_attendance.db'))
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# ----------------- MODELS -----------------

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), unique=True, nullable=False) # e.g., EMP001
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), nullable=True, unique=True)
    role = db.Column(db.String(20), nullable=False) # 'admin', 'faculty', 'student'
    is_active = db.Column(db.Boolean, default=True) # For admin deactivation
    # Adjusted to store a LIST of encodings now
    face_encoding = db.Column(db.PickleType, nullable=True)
    face_registered_at = db.Column(db.DateTime, nullable=True) # When face was registered
    registration_date = db.Column(db.DateTime, default=datetime.utcnow)
    last_active = db.Column(db.DateTime, nullable=True) # For tracking admin activity
    password_hash = db.Column(db.String(128), nullable=True) # Hashed
    # Faculty Registration Approval Workflow
    registration_status = db.Column(db.String(20), default='Approved')  # 'Pending', 'Approved', 'Rejected' (Admins auto-approved, Faculties default pending)
    registration_notes = db.Column(db.String(300), nullable=True)  # Admin feedback on rejection
    registration_submitted_at = db.Column(db.DateTime, nullable=True)  # When faculty submitted self-registration
    # Helper to check password
    def check_password(self, password):
        if not self.password_hash: return False
        return bcrypt.checkpw(password.encode('utf-8'), self.password_hash.encode('utf-8'))

class AttendanceLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    date = db.Column(db.String(20), nullable=False) # Format YYYY-MM-DD
    time_in = db.Column(db.String(20), nullable=True) # Format HH:MM:SS (Local)
    time_out = db.Column(db.String(20), nullable=True) # Format HH:MM:SS (Local)
    check_in_status = db.Column(db.String(30), nullable=True) # P/LP/Absent/HD/EP
    check_out_status = db.Column(db.String(30), nullable=True) # FD/HD/etc
    check_in_period = db.Column(db.String(40), nullable=True)
    check_out_period = db.Column(db.String(40), nullable=True)
    timestamp_in = db.Column(db.DateTime, default=datetime.utcnow) # UTC for cooldown calc
    timestamp_out = db.Column(db.DateTime, nullable=True) # UTC
    status = db.Column(db.String(20), nullable=False) # 'On-Time', 'Late', etc.

class PermissionRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    type = db.Column(db.String(50), nullable=False) # 'custom', 'late_arrival', 'early_departure', 'full_day_absence', 'half_day'
    custom_type = db.Column(db.String(100), nullable=True) # Custom type if type='custom'
    date = db.Column(db.String(20), nullable=False)
    start_time = db.Column(db.String(10), nullable=True) # HH:MM
    end_time = db.Column(db.String(10), nullable=True) # HH:MM
    is_full_day = db.Column(db.Boolean, default=False)
    custom_days = db.Column(db.String(400), nullable=True) # CSV dates: YYYY-MM-DD,YYYY-MM-DD
    reason = db.Column(db.String(500), nullable=True) # Description of the request
    status = db.Column(db.String(20), default='Pending') # 'Pending', 'Approved', 'Rejected'
    document_path = db.Column(db.String(500), nullable=True) # Path to uploaded document/proof
    message_id = db.Column(db.Integer, nullable=True) # Link to chat message if submitted via chat
    approved_by = db.Column(db.String(50), nullable=True) # Admin who approved it
    admin_notes = db.Column(db.String(500), nullable=True) # Notes/conditions from admin
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    admin_id = db.Column(db.String(50), nullable=False)
    action = db.Column(db.String(200), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    reason = db.Column(db.String(200))

class Holiday(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.String(20), unique=True, nullable=False) # YYYY-MM-DD
    name = db.Column(db.String(100), nullable=False)
    type = db.Column(db.String(50), default="Public Holiday")

class LivePresence(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)
    distance_m = db.Column(db.Float, nullable=True)
    in_bounds = db.Column(db.Boolean, default=False)
    status_code = db.Column(db.String(50), default='UNKNOWN')
    status_message = db.Column(db.String(300), nullable=True)
    source = db.Column(db.String(30), default='heartbeat')
    last_seen = db.Column(db.DateTime, default=datetime.utcnow)
    # === NATIVE TRUTH FIELDS (reported by Android service, NOT guessed by JS) ===
    network_status = db.Column(db.String(20), default='online')   # 'online' or 'offline' — from Android ConnectivityManager
    location_enabled = db.Column(db.Boolean, default=True)        # GPS toggle state — from Android LocationManager

class SecurityAlert(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), nullable=True)
    user_name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    event_code = db.Column(db.String(60), nullable=False)
    event_message = db.Column(db.String(300), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class AdminAuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    action = db.Column(db.String(50), nullable=False) # 'admin_created', 'admin_deleted', etc.
    admin_id = db.Column(db.String(50), nullable=False) # Admin who performed the action
    target_id = db.Column(db.String(50), nullable=True) # User being acted upon
    description = db.Column(db.String(300), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class LocationLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    date = db.Column(db.String(20), nullable=False)  # YYYY-MM-DD
    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    distance_m = db.Column(db.Float, nullable=False)  # Distance from target in meters
    in_bounds = db.Column(db.Boolean, nullable=False)  # Whether within allowed radius
    network_status = db.Column(db.String(20), default='online')  # 'online' or 'offline'
    accuracy_m = db.Column(db.Float, nullable=True)  # GPS accuracy in meters
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)  # When the location was captured

    # Index for fast queries
    __table_args__ = (
        db.Index('idx_user_date_timestamp', 'user_id', 'date', 'timestamp'),
    )

class DeviceStateLog(db.Model):
    """Tracks every state TRANSITION for network, GPS location toggle, and geofence bounds.
    Only written when a state actually changes — not on every heartbeat.
    Columns:
      event_type: 'NETWORK' | 'LOCATION' | 'BOUNDS'
      old_state / new_state: describes the transition
        NETWORK:  'online' → 'offline' or 'offline' → 'online'
        LOCATION: 'active' → 'inactive' or 'inactive' → 'active'
        BOUNDS:   'in' → 'out' or 'out' → 'in'
    """
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    date = db.Column(db.String(20), nullable=False)      # YYYY-MM-DD (local)
    event_type = db.Column(db.String(20), nullable=False) # 'NETWORK', 'LOCATION', 'BOUNDS'
    old_state = db.Column(db.String(20), nullable=False)  # Previous state
    new_state = db.Column(db.String(20), nullable=False)  # Current state
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)
    distance_m = db.Column(db.Float, nullable=True)
    accuracy_m = db.Column(db.Float, nullable=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.Index('idx_dsl_user_date', 'user_id', 'date'),
        db.Index('idx_dsl_event_type', 'event_type'),
    )

# NEW FEATURE MODELS

class FCMToken(db.Model):
    """Store Firebase Cloud Messaging tokens for push notifications"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    fcm_token = db.Column(db.String(500), nullable=False, unique=True)
    device_info = db.Column(db.String(200), nullable=True)  # e.g., "Android 12", "iOS"
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Notification(db.Model):
    """Push notifications sent to users"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    type = db.Column(db.String(50), nullable=False)  # 'approval', 'rejection', 'violation', 'announcement', 'alert'
    related_id = db.Column(db.String(100), nullable=True)  # Reference to related entity (e.g., permission_id)
    is_read = db.Column(db.Boolean, default=False)
    is_pinned = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    pinned_at = db.Column(db.DateTime, nullable=True)
    __table_args__ = (
        db.Index('idx_user_created', 'user_id', 'created_at'),
    )

class Message(db.Model):
    """In-app messaging between admins and faculty"""
    id = db.Column(db.Integer, primary_key=True)
    sender_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    recipient_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=True)  # None = broadcast to all
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(20), default='message')  # 'message' or 'announcement'
    is_broadcast = db.Column(db.Boolean, default=False)  # True = send to all faculty
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (
        db.Index('idx_recipient_created', 'recipient_id', 'created_at'),
    )

class MessageRead(db.Model):
    """Track which messages have been read by recipients"""
    id = db.Column(db.Integer, primary_key=True)
    message_id = db.Column(db.Integer, db.ForeignKey('message.id'), nullable=False)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    read_at = db.Column(db.DateTime, nullable=True)
    db.UniqueConstraint('message_id', 'user_id', name='uq_message_user')

class AlertPreference(db.Model):
    """User's notification preferences"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), unique=True, nullable=False)
    # Notification types
    alert_late_arrival = db.Column(db.Boolean, default=True)
    alert_approval_status = db.Column(db.Boolean, default=True)
    alert_policy_violation = db.Column(db.Boolean, default=True)
    alert_announcements = db.Column(db.Boolean, default=True)
    alert_failed_scans = db.Column(db.Boolean, default=True)
    alert_suspicious_activity = db.Column(db.Boolean, default=False)
    # Delivery methods
    enable_in_app_notifications = db.Column(db.Boolean, default=True)
    enable_push_notifications = db.Column(db.Boolean, default=True)
    quiet_hours_start = db.Column(db.String(5), default="22:00")  # HH:MM format
    quiet_hours_end = db.Column(db.String(5), default="08:00")    # HH:MM format
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class DuplicateFaceAlert(db.Model):
    """Alert when duplicate/similar faces are detected during registration"""
    id = db.Column(db.Integer, primary_key=True)
    primary_user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    duplicate_user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    similarity_score = db.Column(db.Float, nullable=False)  # 0-1, higher = more similar
    is_flagged = db.Column(db.Boolean, default=True)
    admin_notes = db.Column(db.String(300), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    resolved_at = db.Column(db.DateTime, nullable=True)

class LivenessChallenge(db.Model):
    """Track liveness check attempts for anti-spoofing"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), nullable=False)
    challenge_type = db.Column(db.String(50), nullable=False)  # 'blink', 'head_turn', 'smile'
    challenge_instructions = db.Column(db.String(300), nullable=False)
    is_passed = db.Column(db.Boolean, nullable=True)  # None = pending, True = passed, False = failed
    attempt_count = db.Column(db.Integer, default=1)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed_at = db.Column(db.DateTime, nullable=True)

class UserThemePreference(db.Model):
    """Store user's theme and UI preferences"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), db.ForeignKey('user.user_id'), unique=True, nullable=False)
    theme = db.Column(db.String(20), default='light')  # 'light', 'dark', 'auto'
    accent_color = db.Column(db.String(7), default='#3b82f6')  # Hex color code
    language = db.Column(db.String(5), default='en')  # Language code
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ----------------- TIME LOGIC HELPER -----------------

ABSENCE_STATUSES = {"Absent", "Didn't Mark"}


def is_absence_status(status):
    return (status or "").strip() in ABSENCE_STATUSES


def has_approved_permission(user_id, date_str, permission_type):
    approved_requests = PermissionRequest.query.filter_by(
        user_id=user_id,
        type=permission_type,
        status='Approved'
    ).all()

    for req in approved_requests:
        if req.date == date_str:
            return True
        if req.custom_days:
            custom_days = [d.strip() for d in req.custom_days.split(',') if d.strip()]
            if date_str in custom_days:
                return True

    return False


def normalize_custom_days(custom_days_raw):
    if custom_days_raw is None:
        return ""

    if isinstance(custom_days_raw, list):
        raw_days = [str(d).strip() for d in custom_days_raw]
    else:
        raw_days = [d.strip() for d in str(custom_days_raw).split(',')]

    valid_days = []
    for day in raw_days:
        if not day:
            continue
        try:
            datetime.strptime(day, '%Y-%m-%d')
            valid_days.append(day)
        except ValueError:
            continue

    return ','.join(sorted(set(valid_days)))


def build_sequential_custom_days(start_date_str, day_count):
    """Build a CSV list of consecutive YYYY-MM-DD dates starting from start_date_str."""
    try:
        count = int(day_count)
    except (TypeError, ValueError):
        return ""

    if count < 1:
        return ""

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except ValueError:
        return ""

    days = [(start_date + timedelta(days=offset)).strftime('%Y-%m-%d') for offset in range(count)]
    return ','.join(days)


def normalize_time_hhmm(raw_time):
    if not raw_time:
        return None
    raw_time = str(raw_time).strip()
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            parsed = datetime.strptime(raw_time, fmt)
            return parsed.strftime('%H:%M')
        except ValueError:
            continue
    return None


def ensure_permission_request_schema():
    """Ensure new PermissionRequest columns exist for older SQLite databases."""
    with db.engine.connect() as conn:
        cols = conn.exec_driver_sql("PRAGMA table_info(permission_request)").fetchall()
        col_names = {c[1] for c in cols}

        # Add missing columns one by one
        if 'custom_type' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN custom_type VARCHAR(100)")
        if 'start_time' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN start_time VARCHAR(10)")
        if 'end_time' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN end_time VARCHAR(10)")
        if 'is_full_day' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN is_full_day BOOLEAN DEFAULT 0")
        if 'custom_days' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN custom_days VARCHAR(400)")
        if 'document_path' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN document_path VARCHAR(500)")
        if 'message_id' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN message_id INTEGER")
        if 'approved_by' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN approved_by VARCHAR(50)")
        if 'admin_notes' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN admin_notes VARCHAR(500)")
        if 'created_at' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
        if 'updated_at' not in col_names:
            conn.exec_driver_sql("ALTER TABLE permission_request ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP")
        
        conn.commit()
        print("✅ Permission Request schema verified/migrated")


def ensure_attendance_log_schema():
    """Ensure new AttendanceLog columns exist for older SQLite databases."""
    with db.engine.connect() as conn:
        cols = conn.exec_driver_sql("PRAGMA table_info(attendance_log)").fetchall()
        col_names = {c[1] for c in cols}

        if 'check_in_status' not in col_names:
            conn.exec_driver_sql("ALTER TABLE attendance_log ADD COLUMN check_in_status VARCHAR(30)")
        if 'check_out_status' not in col_names:
            conn.exec_driver_sql("ALTER TABLE attendance_log ADD COLUMN check_out_status VARCHAR(30)")
        if 'check_in_period' not in col_names:
            conn.exec_driver_sql("ALTER TABLE attendance_log ADD COLUMN check_in_period VARCHAR(40)")
        if 'check_out_period' not in col_names:
            conn.exec_driver_sql("ALTER TABLE attendance_log ADD COLUMN check_out_period VARCHAR(40)")


def ensure_user_schema():
    """Ensure new User columns exist for older SQLite databases."""
    with db.engine.connect() as conn:
        cols = conn.exec_driver_sql("PRAGMA table_info(user)").fetchall()
        col_names = {c[1] for c in cols}

        if 'is_active' not in col_names:
            conn.exec_driver_sql("ALTER TABLE user ADD COLUMN is_active BOOLEAN DEFAULT 1")
        if 'registration_status' not in col_names:
            conn.exec_driver_sql("ALTER TABLE user ADD COLUMN registration_status VARCHAR(20) DEFAULT 'Approved'")
        if 'registration_notes' not in col_names:
            conn.exec_driver_sql("ALTER TABLE user ADD COLUMN registration_notes VARCHAR(500)")
        if 'registration_submitted_at' not in col_names:
            conn.exec_driver_sql("ALTER TABLE user ADD COLUMN registration_submitted_at DATETIME")


def ensure_notification_preferences():
    """Create default alert preferences for users without them."""
    try:
        with app.app_context():
            users_without_prefs = db.session.query(User).filter(
                ~db.session.query(AlertPreference.user_id).filter(
                    AlertPreference.user_id == User.user_id
                ).correlate(User).exists()
            ).all()

            for user in users_without_prefs:
                pref = AlertPreference(user_id=user.user_id)
                db.session.add(pref)

            if users_without_prefs:
                db.session.commit()
                print(f"✓ Created alert preferences for {len(users_without_prefs)} users")
    except Exception as e:
        print(f"[Alert Preference Init Error] {e}")


def ensure_theme_preferences():
    """Create default theme preferences for users without them."""
    try:
        with app.app_context():
            users_without_theme = db.session.query(User).filter(
                ~db.session.query(UserThemePreference.user_id).filter(
                    UserThemePreference.user_id == User.user_id
                ).correlate(User).exists()
            ).all()

            for user in users_without_theme:
                theme_pref = UserThemePreference(user_id=user.user_id)
                db.session.add(theme_pref)

            if users_without_theme:
                db.session.commit()
                print(f"✓ Created theme preferences for {len(users_without_theme)} users")
    except Exception as e:
        print(f"[Theme Preference Init Error] {e}")


def initialize_database_on_startup():
    """Run safe startup initialization for both local runs and WSGI deployments."""
    try:
        with app.app_context():
            db.create_all()
            ensure_user_schema()
            ensure_permission_request_schema()
            ensure_attendance_log_schema()
            ensure_notification_preferences()
            ensure_theme_preferences()

            # Ensure default admin exists for first login
            if not User.query.filter_by(user_id='ADMIN01').first():
                pw_hash = bcrypt.hashpw("admin".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                admin = User(
                    user_id="ADMIN01",
                    name="System Administrators",
                    role="admin",
                    face_encoding=None,
                    password_hash=pw_hash,
                    is_active=True,
                    registration_status='Approved'
                )
                db.session.add(admin)
                db.session.commit()
                print(">>> Default Admin Created: User=ADMIN01, Pass=admin")
    except Exception as e:
        print(f"[Startup Init Error] {e}")


# IMPORTANT: Run startup init on import as well (required for PythonAnywhere WSGI).
initialize_database_on_startup()

# ════════════════════════════════════════════════════════════════════
# 🔴 START FCM BACKGROUND SCHEDULER: Ping devices every 10 minutes
# ════════════════════════════════════════════════════════════════════
if APSCHEDULER_AVAILABLE:
    try:
        print("\n" + "="*60)
        print("🔴 [FCM SCHEDULER] Initializing background scheduler...")
        print("="*60)
        
        scheduler = BackgroundScheduler(daemon=True)
        print("✅ BackgroundScheduler object created")
        
        # Add background job: ping all active devices every 10 minutes
        scheduler.add_job(
            func=ping_all_active_devices,
            trigger=IntervalTrigger(minutes=10),
            id='fcm_device_ping',
            name='FCM Device Ping (every 10 min)',
            replace_existing=True,
            max_instances=1  # Only one instance at a time
        )
        print("✅ Job 1 added: FCM Device Pings (every 10 min)")
        
        # Add background job: ping inactive devices every 5 minutes (more aggressive)
        scheduler.add_job(
            func=ping_inactive_devices,
            trigger=IntervalTrigger(minutes=5),
            id='fcm_inactive_ping',
            name='FCM Inactive Device Wakeup (every 5 min)',
            replace_existing=True,
            max_instances=1
        )
        print("✅ Job 2 added: Inactive Device Wakeup (every 5 min)")
        
        # Add background job: auto-mark incomplete attendance daily at 6:00 PM
        try:
            from apscheduler.triggers.cron import CronTrigger
            scheduler.add_job(
                func=auto_mark_incomplete_attendance,
                trigger=CronTrigger(hour=18, minute=0, timezone='Asia/Kolkata'),
                id='auto_mark_attendance',
                name='Auto-Mark Incomplete Attendance (6:00 PM daily)',
                replace_existing=True,
                max_instances=1
            )
            print("✅ Job 2 added: Auto-Mark Absent (6:00 PM IST daily)")
        except ImportError:
            print("⚠️ CronTrigger not available - auto-mark scheduled job DISABLED")
            print("   Use /api/mark/auto-mark-absent endpoint to trigger manually")
        
        scheduler.start()
        print("✅ SCHEDULER STARTED! Now running in background...")
        print("   - FCM Device Pings: every 10 minutes (all logged-in users)")
        print("   - Inactive Device Wakeup: every 5 minutes (devices with no updates)")
        print("   - Auto-Mark Absent: 6:00 PM IST daily")
        print("="*60 + "\n")
    except Exception as e:
        print(f"\n❌ [SCHEDULER ERROR] FCM Scheduler FAILED to start: {e}")
        print("   Devices won't be pinged periodically. Location tracking may not work reliably.")
        import traceback
        traceback.print_exc()
        print()
        scheduler = None
else:
    print("\n" + "="*60)
    print("⚠️ [SCHEDULER DISABLED] APScheduler not available")
    print("   Devices will only report location when they detect a change (manual heartbeat)")
    print("   Install apscheduler to enable periodic FCM pings: pip install apscheduler")
    print("="*60 + "\n")
    scheduler = None  # No scheduler available

def require_active_admin(admin_id):
    if not admin_id:
        return None
    return User.query.filter_by(user_id=admin_id, role='admin', is_active=True).first()


@app.route('/api/mark/auto-mark-absent', methods=['POST'])
def trigger_auto_mark_absent():
    """
    🔴 MANUAL TRIGGER: Auto-mark faculty as ABSENT if no 2nd mark by 6 PM.
    
    Call this endpoint at 6:00 PM to complete end-of-day attendance marking.
    Useful for PythonAnywhere if APScheduler is not available.
    
    Requires admin authentication.
    """
    data = request.json or {}
    admin_id = (data.get('admin_id') or '').strip()
    
    admin_user = require_active_admin(admin_id)
    if not admin_user:
        return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403
    
    try:
        auto_mark_incomplete_attendance()
        return jsonify({
            "success": True,
            "message": "Auto-mark attendance job triggered successfully"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error triggering auto-mark: {str(e)}"
        }), 500


def build_permission_summary(permission, status_label=None, admin_notes=None):
    permission_label = permission.custom_type if permission.type == 'custom' and permission.custom_type else permission.type.replace('_', ' ').title()
    lines = [
        "📋 Permission Request",
        f"Type: {permission_label}",
        f"Date: {permission.date}"
    ]

    if permission.start_time or permission.end_time:
        lines.append(f"Time: {(permission.start_time or '--:--')} to {(permission.end_time or '--:--')}")

    if permission.is_full_day:
        lines.append("Request: Full day")

    if permission.reason:
        lines.append(f"Reason: {permission.reason}")

    if status_label:
        lines.append(f"Status: {status_label}")

    if admin_notes:
        lines.append(f"Admin notes: {admin_notes}")

    return "\n".join(lines)


def create_permission_request_chat_message(permission, recipient_id):
    message = Message(
        sender_id=permission.user_id,
        recipient_id=recipient_id,
        title='Permission Request',
        content=build_permission_summary(permission, status_label='Pending'),
        type='permission_request',
        is_broadcast=False
    )
    db.session.add(message)
    db.session.flush()

    if recipient_id:
        db.session.add(MessageRead(
            message_id=message.id,
            user_id=recipient_id,
            is_read=False
        ))

    return message


def create_permission_decision_chat_message(permission, sender_id, decision, admin_notes=None):
    recipient_id = permission.user_id
    message = Message(
        sender_id=sender_id,
        recipient_id=recipient_id,
        title=f'Permission {decision}',
        content=build_permission_summary(permission, status_label=decision, admin_notes=admin_notes),
        type='message',
        is_broadcast=False
    )
    db.session.add(message)
    db.session.flush()
    db.session.add(MessageRead(
        message_id=message.id,
        user_id=recipient_id,
        is_read=False
    ))
    return message


@app.route('/api/permissions/request', methods=['POST'])
def submit_permission_request():
    """Faculty submits a permission request via chat or directly."""
    try:
        user_id = (request.form.get('user_id') or '').strip()
        recipient_id = (request.form.get('recipient_id') or '').strip()  # Admin receiving the request
        perm_type = (request.form.get('type') or '').strip()
        custom_type = (request.form.get('custom_type') or '').strip()
        date_str = (request.form.get('date') or '').strip()
        custom_days_raw = request.form.get('custom_days')
        custom_days_count_raw = (request.form.get('custom_days_count') or '').strip()
        start_time = (request.form.get('start_time') or '').strip()
        end_time = (request.form.get('end_time') or '').strip()
        is_full_day = request.form.get('is_full_day', 'false').lower() == 'true'
        reason = (request.form.get('reason') or '').strip()

        # Validate required fields
        if not user_id or not perm_type or not date_str or not reason:
            return jsonify({"success": False, "message": "user_id, type, date, and reason are required."}), 400

        # Validate date format
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({"success": False, "message": "date must be in YYYY-MM-DD format."}), 400

        custom_days = normalize_custom_days(custom_days_raw)
        if perm_type == 'custom' and not custom_days:
            custom_days = build_sequential_custom_days(date_str, custom_days_count_raw or 1)
            if not custom_days:
                return jsonify({"success": False, "message": "custom_days_count must be a positive number."}), 400

        # Validate faculty exists
        faculty = User.query.filter_by(user_id=user_id).first()
        if not faculty:
            return jsonify({"success": False, "message": "Faculty not found."}), 404

        if faculty.role != 'faculty':
            return jsonify({"success": False, "message": "Only faculty can raise permission requests."}), 403

        if not recipient_id:
            return jsonify({"success": False, "message": "recipient_id is required and must be an admin."}), 400

        recipient = User.query.filter_by(user_id=recipient_id).first()
        if not recipient or recipient.role != 'admin':
            return jsonify({"success": False, "message": "Permission requests can only be sent to admins."}), 400

        # Handle time validation for partial-day permissions
        if not is_full_day and perm_type in ['late_arrival', 'early_departure', 'half_day']:
            if not start_time or not end_time:
                return jsonify({"success": False, "message": "start_time and end_time are required for partial-day permissions."}), 400
            
            try:
                start_obj = datetime.strptime(start_time, '%H:%M')
                end_obj = datetime.strptime(end_time, '%H:%M')
                if end_obj <= start_obj:
                    return jsonify({"success": False, "message": "end_time must be later than start_time."}), 400
            except ValueError:
                return jsonify({"success": False, "message": "Times must be in HH:MM format."}), 400

        # Handle file upload
        document_path = None
        if 'document' in request.files:
            file = request.files['document']
            if file and file.filename:
                # Check file size (5MB limit)
                file.seek(0, 2)
                file_size = file.tell()
                file.seek(0)
                
                if file_size > 5 * 1024 * 1024:
                    return jsonify({"success": False, "message": "File size must not exceed 5MB."}), 400

                # Save file
                allowed_extensions = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png'}
                file_ext = file.filename.split('.')[-1].lower()
                
                if file_ext not in allowed_extensions:
                    return jsonify({"success": False, "message": f"File type .{file_ext} not allowed."}), 400

                filename = f"perm_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_ext}"
                uploads_dir = os.path.join(os.path.dirname(__file__), 'uploads', 'permissions')
                os.makedirs(uploads_dir, exist_ok=True)
                
                file_path = os.path.join(uploads_dir, filename)
                file.save(file_path)
                document_path = f"uploads/permissions/{filename}"

        # Create permission request
        permission = PermissionRequest(
            user_id=user_id,
            type=perm_type,
            custom_type=custom_type if perm_type == 'custom' else None,
            date=date_str,
            start_time=start_time if not is_full_day else None,
            end_time=end_time if not is_full_day else None,
            is_full_day=is_full_day,
            custom_days=custom_days if perm_type == 'custom' else None,
            reason=reason,
            document_path=document_path,
            status='Pending'
        )

        db.session.add(permission)
        db.session.flush()

        chat_message = create_permission_request_chat_message(permission, recipient_id)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Permission request submitted successfully!",
            "permission_id": permission.id,
            "message_id": chat_message.id,
            "permission": {
                "id": permission.id,
                "user_id": permission.user_id,
                "type": permission.type,
                "custom_type": permission.custom_type,
                "date": permission.date,
                "start_time": permission.start_time,
                "end_time": permission.end_time,
                "is_full_day": permission.is_full_day,
                "custom_days": permission.custom_days,
                "reason": permission.reason,
                "status": permission.status,
                "created_at": permission.created_at.isoformat() if permission.created_at else None
            }
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"Error in submit_permission_request: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


def classify_first_mark(scan_dt):
    """Return first-mark status for the day based on configured time windows."""
    t = scan_dt.time()

    if t < time(9, 0):
        return {
            "allowed": False,
            "status": None,
            "period": "00:00-09:00",
            "message": "Attendance is not taken between 12:00 AM and 9:00 AM."
        }
    if t < time(9, 35):
        return {"allowed": True, "status": "Present", "period": "09:00-09:35", "message": "Check-In: Present"}
    if t < time(10, 30):
        return {"allowed": True, "status": "Late Permission", "period": "09:35-10:30", "message": "Check-In: Late Permission"}
    if t < time(12, 40):
        return {"allowed": True, "status": "Absent", "period": "10:30-12:40", "message": "Check-In: Absent"}
    if t < time(13, 40):
        return {"allowed": True, "status": "HD", "period": "12:40-13:40", "message": "Check-In: Half Day"}
    if t < time(15, 10):
        return {"allowed": True, "status": "Absent", "period": "13:40-15:10", "message": "Check-In: Absent"}
    if t < time(16, 10):
        return {"allowed": True, "status": "EP", "period": "15:10-16:10", "message": "Check-In: Early Permission"}

    return {
        "allowed": False,
        "status": None,
        "period": "16:10-24:00",
        "message": "First marking is closed in evening window. Evening marking is only for second mark."
    }


def classify_second_mark(first_status, scan_dt):
    """Return second-mark final status for the day based on first status and time window."""
    t = scan_dt.time()
    first_status = (first_status or "").strip()

    if time(12, 40) <= t < time(13, 40) and first_status in {"Present", "Late Permission"}:
        return {
            "allowed": True,
            "final_status": "HD",
            "out_status": "HD",
            "period": "12:40-13:40",
            "message": "Check-Out: Half Day"
        }

    if time(16, 10) <= t:
        if first_status in {"Present", "Late Permission"}:
            return {
                "allowed": True,
                "final_status": "FD",
                "out_status": "FD",
                "period": "16:10-24:00",
                "message": "Check-Out: Full Day"
            }

        if first_status in {"Absent", "HD", "EP"}:
            return {
                "allowed": True,
                "final_status": "HD",
                "out_status": "HD",
                "period": "16:10-24:00",
                "message": "Check-Out: Half Day"
            }

        return {
            "allowed": True,
            "final_status": "HD",
            "out_status": "HD",
            "period": "16:10-24:00",
            "message": "Check-Out: Half Day"
        }

    return {
        "allowed": False,
        "final_status": None,
        "out_status": None,
        "period": None,
        "message": "Second mark ignored in this period."
    }

def check_attendance_status(user_id, scan_time, type, late_permission_approved=False):
    """
    Check-In:
      < 09:30: On Time
      09:30 - 09:45: Late Permission
      > 09:45: Absent unless a late permission has been approved

    Check-Out:
      < 16:10: Early Permission
      16:10 - 17:00: On Time
      > 17:00: Extended
    """

    # Define constraints
    COLLEGE_START = time(9, 30, 0)
    GRACE_END = time(9, 45, 0)

    COLLEGE_END = time(16, 10, 0) # 4:10 PM
    EXTENDED_END = time(17, 0, 0) # 5:00 PM

    today_str = scan_time.strftime('%Y-%m-%d')
    scan_time_val = scan_time.time()

    if type == 'IN':
        if scan_time_val < COLLEGE_START:
            return "On Time"
        elif scan_time_val <= GRACE_END:
            return "Late Permission"
        elif late_permission_approved:
            return "Late Permission"
        else:
            return "Absent"

    elif type == 'OUT':
        if scan_time_val < COLLEGE_END:
             return "Early Permission"
        elif scan_time_val <= EXTENDED_END:
             return "On Time"
        else:
             return "Extended"

    return "Unknown"


def has_approved_permission(user_id, date_str, permission_type=None):
    """Check if faculty has an approved permission for the given date."""
    try:
        query = PermissionRequest.query.filter_by(
            user_id=user_id,
            date=date_str,
            status='Approved'
        )
        
        if permission_type:
            query = query.filter(PermissionRequest.type.in_([permission_type, 'custom']))
        
        permission = query.first()
        return bool(permission)
    except Exception as e:
        print(f"Error checking permission: {e}")
        return False


def get_time_policy_flags(local_dt):
    """Return time-window flags used by frontend monitoring and backend enforcement."""
    now_t = local_dt.time()
    return {
        "is_check_in_window": time(9, 0, 0) <= now_t <= time(9, 45, 0),
        "is_lunch_window": time(13, 0, 0) <= now_t <= time(13, 40, 0),
        "is_lunch_pre_alert": time(12, 50, 0) <= now_t < time(13, 0, 0),
        "is_return_pre_alert": time(13, 30, 0) <= now_t < time(13, 40, 0),
        "is_post_lunch_enforcement": time(13, 40, 0) < now_t < time(16, 10, 0),
        "is_check_out_window": time(16, 10, 0) <= now_t <= time(17, 0, 0)
    }


def upsert_live_presence(user, status_code, status_message, source='heartbeat', latitude=None, longitude=None, distance_m=None, in_bounds=False, native_network_status=None, native_location_enabled=None):
    """Create or update current live presence state for a user.
    
    native_network_status: 'online' or 'offline' — ONLY set by native Android service
    native_location_enabled: True/False — ONLY set by native Android service
    """
    presence = LivePresence.query.filter_by(user_id=user.user_id).first()
    previous_status = None

    if not presence:
        presence = LivePresence(
            user_id=user.user_id,
            name=user.name,
            role=user.role
        )
        db.session.add(presence)
    else:
        previous_status = presence.status_code

    presence.name = user.name
    presence.role = user.role

    # Store native truth fields if provided (ONLY from native Android service)
    if native_network_status is not None:
        presence.network_status = native_network_status
    if native_location_enabled is not None:
        presence.location_enabled = bool(native_location_enabled)

    # PREVENT WEB APP FROM OVERWRITING NATIVE FAULTS
    if source != 'native_tracker' and previous_status in ['NETWORK_OFF', 'LOCATION_OFF']:
        # If native service previously declared a fault, but web app is trying to say 'OK'
        # based on cached browser data, ignore the web app's OK status.
        if presence.network_status == 'offline':
            status_code = 'NETWORK_OFF'
            status_message = 'Device network disconnected (kept native)'
        elif not presence.location_enabled:
            status_code = 'LOCATION_OFF'
            status_message = 'GPS disabled on device (kept native)'

    presence.status_code = status_code
    presence.status_message = status_message
    presence.source = source
    presence.last_seen = datetime.utcnow()
    presence.in_bounds = bool(in_bounds)
    presence.distance_m = distance_m

    if latitude is not None and longitude is not None:
        presence.latitude = latitude
        presence.longitude = longitude

    # Create alert if status is not OK (either new alert or state change)
    # OUT_OF_BOUNDS alerts are created every time to ensure admin always sees current violations
    should_create_alert = (status_code != 'OK') and (previous_status != status_code or status_code == 'OUT_OF_BOUNDS')

    if should_create_alert:
        existing_alert = SecurityAlert.query.filter_by(
            user_id=user.user_id,
            event_code=status_code
        ).filter(
            SecurityAlert.created_at >= (datetime.utcnow() - timedelta(minutes=5))
        ).first()

        if not existing_alert:
            db.session.add(SecurityAlert(
                user_id=user.user_id,
                user_name=user.name,
                role=user.role,
                event_code=status_code,
                event_message=status_message
            ))
            print(f"DEBUG: SecurityAlert created for {user.user_id}: {status_code} - {status_message}")

def verify_presence_by_location(user_id, date_str, min_location_entries=3):
    """
    Verify if a user has continuous location tracking for a date.
    Returns True if user has at least min_location_entries valid location pings.

    This ensures that even if cache is cleared or app crashes,
    the location data proves they were present on campus.
    """
    try:
        location_logs = LocationLog.query.filter_by(
            user_id=user_id,
            date=date_str,
            in_bounds=True
        ).count()

        # If at least 3 valid in-bounds location pings, consider them present
        return location_logs >= min_location_entries
    except:
        return False

def get_user_location_history(user_id, date_str):
    """
    Get all location pings for a user on a specific date.
    Used for reporting and verification.
    """
    try:
        logs = LocationLog.query.filter_by(
            user_id=user_id,
            date=date_str
        ).order_by(LocationLog.timestamp.asc()).all()

        return [{
            'timestamp': log.timestamp.isoformat(),
            'latitude': log.latitude,
            'longitude': log.longitude,
            'distance_m': log.distance_m,
            'in_bounds': log.in_bounds,
            'accuracy_m': log.accuracy_m
        } for log in logs]
    except:
        return []

def ensure_location_log_table():
    """
    Ensure LocationLog table exists and has required indexes.
    Called when app starts.
    """
    try:
        with app.app_context():
            # Create table if it doesn't exist
            db.create_all()
        print("[DB] LocationLog table verified")
    except Exception as e:
        print(f"[DB ERROR] LocationLog setup failed: {e}")

def ensure_live_presence_schema():
    """Ensure new LivePresence columns exist for older SQLite databases."""
    try:
        with app.app_context():
            conn = db.engine.raw_connection()
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(live_presence)")
            cols = cursor.fetchall()
            col_names = {c[1] for c in cols}

            if 'network_status' not in col_names:
                cursor.execute("ALTER TABLE live_presence ADD COLUMN network_status VARCHAR(20) DEFAULT 'online'")
                print("[DB] Added network_status column to live_presence")
            if 'location_enabled' not in col_names:
                cursor.execute("ALTER TABLE live_presence ADD COLUMN location_enabled BOOLEAN DEFAULT 1")
                print("[DB] Added location_enabled column to live_presence")
            
            conn.commit()
            cursor.close()
            conn.close()
    except Exception as e:
        print(f"[DB] LivePresence schema check: {e}")

# Call this on app startup
ensure_location_log_table()
ensure_live_presence_schema()

# ----------------- ROUTES -----------------

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('.', path)

# 1. Initialize Database
@app.route('/api/init_db', methods=['POST'])
def init_db():
    try:
        db.drop_all() # Reset
        db.create_all()

        # Create Default Admin
        pw_hash = bcrypt.hashpw("admin".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        admin = User(
            user_id="ADMIN01",
            name="System Administrator",
            role="admin",
            face_encoding=None,
            password_hash=pw_hash
        )
        db.session.add(admin)
        db.session.commit()

        return jsonify({"message": "Database initialized. Admin created (ID: ADMIN01, Pass: admin)"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# 2. Register User (New or Add Face)
# 2. Register User (New or Add Face)
@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    try:
        # data: name, user_id, role, image (base64). Optional: password (for auth)
        print(f"Registering for {data['user_id']}")

        # Decode image
        image_data = base64.b64decode(data['image'].split(',')[1])
        image_bytes = image_data

        # Get face encoding and location
        encoding, loc = face_system.get_face_encoding(image_bytes)

        if encoding is None:
            return jsonify({"success": False, "message": "Face quality check failed or no face detected."}), 400

        # Check if user exists
        existing_user = User.query.filter_by(user_id=data['user_id']).first()

        if existing_user:
            # Mode: Update Face (Retraining)
            raw_pw = data.get('password')
            if raw_pw and existing_user.check_password(raw_pw):
                current_encodings = existing_user.face_encoding
                if current_encodings is None:
                    current_encodings = []
                elif not isinstance(current_encodings, list):
                    current_encodings = [current_encodings]

                new_encodings_list = list(current_encodings)
                new_encodings_list.append(encoding)

                existing_user.face_encoding = new_encodings_list
                db.session.commit()
                return jsonify({"success": True, "message": "Face model updated successfully."})
            else:
                 return jsonify({"success": False, "message": "User exists. Password required to update face."}), 401

        else:
            # Mode: New User
            raw_pw = data.get('password', data['user_id'])
            pw_hash = bcrypt.hashpw(raw_pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

            new_user = User(
                user_id=data['user_id'],
                name=data['name'],
                role=data['role'],
                password_hash=pw_hash,
                face_encoding=[encoding] # Start list
            )
            db.session.add(new_user)
            db.session.commit()
            return jsonify({"success": True, "message": "User registered successfully"})

    except Exception as e:
        print(f"Registration Error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# Simple health check endpoint (no database needed)
@app.route('/api/health', methods=['GET', 'POST', 'OPTIONS'])
def health_check():
    """Endpoint to verify backend is reachable and diagnostics"""
    print("[Health] Connection test from:", request.remote_addr)

    face_system_status = {
        "models_loaded": face_system.models_loaded,
        "detector_path": face_system.detector_path if hasattr(face_system, 'detector_path') else "N/A",
        "recognizer_path": face_system.recognizer_path if hasattr(face_system, 'recognizer_path') else "N/A",
    }

    # Check if model files exist
    if hasattr(face_system, 'detector_path') and os.path.exists(face_system.detector_path):
        face_system_status["detector_exists"] = True
        face_system_status["detector_size_mb"] = round(os.path.getsize(face_system.detector_path) / (1024*1024), 2)
    else:
        face_system_status["detector_exists"] = False

    if hasattr(face_system, 'recognizer_path') and os.path.exists(face_system.recognizer_path):
        face_system_status["recognizer_exists"] = True
        face_system_status["recognizer_size_mb"] = round(os.path.getsize(face_system.recognizer_path) / (1024*1024), 2)
    else:
        face_system_status["recognizer_exists"] = False

    return jsonify({
        "status": "ok" if face_system.models_loaded else "warning",
        "message": "Backend is running",
        "face_system": face_system_status,
        "debug_info": {
            "timestamp": datetime.now(pytz.utc).isoformat(),
            "location_enforcement": LOCATION_ENFORCEMENT_ENABLED
        }
    }), 200

# ════════════════════════════════════════════════════════════════════
# 🔴 DEVICE STATE LOGGING HELPERS - Accurate state tracking at every login/logout
# ════════════════════════════════════════════════════════════════════

def log_initial_device_states(user_id):
    """
    🔴 CRITICAL: Called at LOGIN to establish baseline device states.
    Creates initial state entries for NETWORK, LOCATION, and BOUNDS.
    All subsequent state changes are compared against these baselines.
    
    This ensures accurate logging of:
    - Login time (baseline recorded)
    - First network state change
    - First GPS toggle change
    - First geofence boundary change
    """
    try:
        now_utc = datetime.now(pytz.utc)
        local_tz = pytz.timezone('Asia/Kolkata')
        now_local = now_utc.astimezone(local_tz)
        today_str = now_local.strftime('%Y-%m-%d')
        
        # ✅ Create baseline state entries at login (all starting as active)
        # These will be the reference point for future state change comparisons
        
        # 1. NETWORK baseline: assume 'online' when user logs in
        existing_net = DeviceStateLog.query.filter_by(
            user_id=user_id, date=today_str, event_type='NETWORK'
        ).first()
        if not existing_net:
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='NETWORK',
                old_state='offline', new_state='online',  # Baseline: online at login
                latitude=None, longitude=None, distance_m=None,
                accuracy_m=None, timestamp=now_utc
            ))
            print(f"[StateLog] {user_id} LOGIN: NETWORK baseline → online")
        
        # 2. LOCATION baseline: assume 'active' when user logs in (GPS assumed enabled)
        existing_loc = DeviceStateLog.query.filter_by(
            user_id=user_id, date=today_str, event_type='LOCATION'
        ).first()
        if not existing_loc:
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='LOCATION',
                old_state='inactive', new_state='active',  # Baseline: GPS active at login
                latitude=None, longitude=None, distance_m=None,
                accuracy_m=None, timestamp=now_utc
            ))
            print(f"[StateLog] {user_id} LOGIN: LOCATION baseline → active")
        
        # 3. BOUNDS baseline: assume 'in' when user logs in (reasonable assumption)
        existing_bnd = DeviceStateLog.query.filter_by(
            user_id=user_id, date=today_str, event_type='BOUNDS'
        ).first()
        if not existing_bnd:
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='BOUNDS',
                old_state='out', new_state='in',  # Baseline: in bounds at login
                latitude=None, longitude=None, distance_m=None,
                accuracy_m=None, timestamp=now_utc
            ))
            print(f"[StateLog] {user_id} LOGIN: BOUNDS baseline → in")
        
        db.session.commit()
        return True
    except Exception as e:
        print(f"[StateLog ERROR] Failed to log initial states for {user_id}: {str(e)}")
        db.session.rollback()
        return False

def log_final_device_states(user_id):
    """
    🔴 CRITICAL: Called at LOGOUT to finalize device states.
    Transitions NETWORK→offline, LOCATION→inactive, BOUNDS→out.
    This creates the final state record for accurate session tracking.
    """
    try:
        now_utc = datetime.now(pytz.utc)
        local_tz = pytz.timezone('Asia/Kolkata')
        now_local = now_utc.astimezone(local_tz)
        today_str = now_local.strftime('%Y-%m-%d')
        
        def get_last_device_state(evt_type):
            return DeviceStateLog.query.filter_by(
                user_id=user_id, date=today_str, event_type=evt_type
            ).order_by(DeviceStateLog.timestamp.desc()).first()
        
        # ✅ Log final state transitions at logout
        
        # 1. NETWORK: online → offline
        last_net = get_last_device_state('NETWORK')
        prev_net = last_net.new_state if last_net else 'online'
        if prev_net != 'offline':
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='NETWORK',
                old_state=prev_net, new_state='offline',
                latitude=None, longitude=None, distance_m=None,
                accuracy_m=None, timestamp=now_utc
            ))
            print(f"[StateLog] {user_id} LOGOUT: NETWORK {prev_net} → offline")
        
        # 2. LOCATION: active → inactive
        last_loc = get_last_device_state('LOCATION')
        prev_loc = last_loc.new_state if last_loc else 'active'
        if prev_loc != 'inactive':
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='LOCATION',
                old_state=prev_loc, new_state='inactive',
                latitude=None, longitude=None, distance_m=None,
                accuracy_m=None, timestamp=now_utc
            ))
            print(f"[StateLog] {user_id} LOGOUT: LOCATION {prev_loc} → inactive")
        
        # 3. BOUNDS: in/out → out
        last_bnd = get_last_device_state('BOUNDS')
        prev_bnd = last_bnd.new_state if last_bnd else 'in'
        if prev_bnd != 'out':
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='BOUNDS',
                old_state=prev_bnd, new_state='out',
                latitude=None, longitude=None, distance_m=None,
                accuracy_m=None, timestamp=now_utc
            ))
            print(f"[StateLog] {user_id} LOGOUT: BOUNDS {prev_bnd} → out")
        
        db.session.commit()
        return True
    except Exception as e:
        print(f"[StateLog ERROR] Failed to log final states for {user_id}: {str(e)}")
        db.session.rollback()
        return False

# 3. Login (Credentials) - Auto-detect role from database
@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json(silent=True) or {}
        username = (data.get('username') or '').strip()
        password = data.get('password') or ''

        if not username or not password:
            return jsonify({"success": False, "message": "Username and password are required"}), 400

        user = User.query.filter_by(user_id=username).first()

        if not user or not user.check_password(password):
            return jsonify({"success": False, "message": "Invalid Credentials"}), 401

        # Check if admin account is deactivated
        if user.role == 'admin' and not user.is_active:
            return jsonify({"success": False, "message": "Your admin account has been deactivated. Contact system administrator."}), 403

        # Check if faculty registration is pending or rejected
        if user.role == 'faculty' and user.registration_status == 'Pending':
            return jsonify({
                "success": False,
                "message": "Your registration is pending admin approval. Please wait for approval notification."
            }), 403

        if user.role == 'faculty' and user.registration_status == 'Rejected':
            reason = user.registration_notes or "No reason provided"
            return jsonify({
                "success": False,
                "message": f"Your registration was rejected. Reason: {reason}"
            }), 403

        # Check if user account is deactivated
        if not user.is_active:
            return jsonify({"success": False, "message": "Your account is deactivated. Contact admin."}), 403

        # Check if user needs to complete face registration (first login)
        needs_face = False
        if user.role == 'faculty' and not user.face_encoding:
            needs_face = True
        elif user.role == 'admin' and not user.face_encoding and user.user_id != 'ADMIN01':
            needs_face = True

        # 🔴 LOG INITIAL DEVICE STATES AT LOGIN: Establish baseline for state tracking
        log_initial_device_states(user.user_id)

        return jsonify({
            "success": True,
            "user": {
                "id": user.user_id,
                "user_id": user.user_id,
                "name": user.name,
                "role": user.role
            },
            "needs_face_registration": needs_face
        })
    except Exception as e:
        print(f"Login Error: {e}")
        return jsonify({"success": False, "message": "Internal server error during login"}), 500

# 3a. Logout - Clear user's live presence and tracking
@app.route('/api/logout', methods=['POST'])
def logout():
    """Clean up user's live presence data on logout and record device states"""
    try:
        data = request.json or {}
        user_id = (data.get('user_id') or '').strip()

        if not user_id:
            return jsonify({"success": False, "message": "user_id is required"}), 400

        # 🔴 LOG FINAL DEVICE STATES AT LOGOUT: Transition to offline/inactive/out
        log_final_device_states(user_id)

        # Delete user's LivePresence record
        presence = LivePresence.query.filter_by(user_id=user_id).first()
        if presence:
            db.session.delete(presence)
            db.session.commit()
            print(f"✅ Logout: LivePresence cleared for {user_id}")
        else:
            print(f"✅ Logout: No LivePresence record found for {user_id} (already cleaned)")

        return jsonify({
            "success": True,
            "message": f"User {user_id} logged out. Tracking stopped."
        })

    except Exception as e:
        print(f"Logout Error: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

# 3a-cleanup. Auto-cleanup stale presence records
def cleanup_stale_presence():
    """Remove LivePresence records older than 1 hour.
    NOTE: Network/Location status is now computed dynamically in admin_live_locations
    based on last_seen freshness, so we don't need to mutate records here.
    Also logs state transitions for stale users (offline/inactive/out).
    """
    try:
        now_utc = datetime.utcnow()

        # Remove records older than 1 hour (user likely logged out or phone died long ago)
        stale_delete_cutoff = now_utc - timedelta(hours=1)
        stale_records = LivePresence.query.filter(
            LivePresence.last_seen < stale_delete_cutoff
        ).all()

        count = len(stale_records)
        if count > 0:
            local_tz = pytz.timezone('Asia/Kolkata')
            today_str = datetime.now(pytz.utc).astimezone(local_tz).strftime('%Y-%m-%d')
            
            for record in stale_records:
                user_id = record.user_id
                
                # 🔴 Log all three state transitions for stale users (cleanup implies they're inactive)
                def get_last_device_state(evt_type):
                    return DeviceStateLog.query.filter_by(
                        user_id=user_id, date=today_str, event_type=evt_type
                    ).order_by(DeviceStateLog.timestamp.desc()).first()
                
                # Log NETWORK: online → offline
                last_net = get_last_device_state('NETWORK')
                prev_net = last_net.new_state if last_net else 'online'
                if prev_net != 'offline':
                    db.session.add(DeviceStateLog(
                        user_id=user_id, date=today_str, event_type='NETWORK',
                        old_state=prev_net, new_state='offline',
                        latitude=record.latitude, longitude=record.longitude,
                        distance_m=record.distance_m, accuracy_m=0.0, timestamp=now_utc
                    ))
                
                # Log LOCATION: active → inactive
                last_loc = get_last_device_state('LOCATION')
                prev_loc = last_loc.new_state if last_loc else 'active'
                if prev_loc != 'inactive':
                    db.session.add(DeviceStateLog(
                        user_id=user_id, date=today_str, event_type='LOCATION',
                        old_state=prev_loc, new_state='inactive',
                        latitude=record.latitude, longitude=record.longitude,
                        distance_m=record.distance_m, accuracy_m=0.0, timestamp=now_utc
                    ))
                
                # Log BOUNDS: in → out
                last_bnd = get_last_device_state('BOUNDS')
                prev_bnd = last_bnd.new_state if last_bnd else 'in'
                if prev_bnd != 'out':
                    db.session.add(DeviceStateLog(
                        user_id=user_id, date=today_str, event_type='BOUNDS',
                        old_state=prev_bnd, new_state='out',
                        latitude=record.latitude, longitude=record.longitude,
                        distance_m=record.distance_m, accuracy_m=0.0, timestamp=now_utc
                    ))
                
                db.session.delete(record)
            
            db.session.commit()
            print(f"🧹 Cleanup: Removed {count} stale LivePresence records older than 1 hour (state transitions logged)")
    except Exception as e:
        print(f"Cleanup Error: {e}")
        db.session.rollback()

# 3b. Register Faculty (Admin Only)
@app.route('/api/register_faculty', methods=['POST'])
def register_faculty():
    data = request.json
    admin_id = data.get('admin_id')

    try:
        # Verify admin
        admin = User.query.filter_by(user_id=admin_id).first()
        if not admin or admin.role != 'admin':
            return jsonify({"success": False, "message": "Unauthorized. Admin access required."}), 403

        # Check if faculty already exists
        if User.query.filter_by(user_id=data['user_id']).first():
            return jsonify({"success": False, "message": "Faculty ID already exists"}), 400

        # Create faculty user
        pw_hash = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        new_faculty = User(
            user_id=data['user_id'],
            name=data['name'],
            role='faculty',
            password_hash=pw_hash,
            face_encoding=None  # Faculty must register face on first login
        )

        db.session.add(new_faculty)
        db.session.commit()

        # Audit log
        audit = AuditLog(
            admin_id=admin_id,
            action=f"Registered new faculty: {data['name']} ({data['user_id']})",
            timestamp=datetime.now(pytz.utc)
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Faculty '{data['name']}' registered successfully. They must capture face on first login.",
            "faculty": {
                "user_id": new_faculty.user_id,
                "name": new_faculty.name,
                "role": new_faculty.role
            }
        })

    except Exception as e:
        print(f"Faculty Registration Error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ============================================
# ADMIN MANAGEMENT ENDPOINTS
# ============================================

# 3B. Create Admin (Direct Registration)
@app.route('/api/create_admin', methods=['POST'])
def create_admin():
    """Admin creates another admin directly"""
    try:
        # Verify admin creating the request
        current_user_id = request.json.get('created_by') or 'ADMIN01'  # From auth header in production
        admin_user = User.query.filter_by(user_id=current_user_id, role='admin', is_active=True).first()

        if not admin_user:
            return jsonify({"error": "Unauthorized. Only active admins can create new admins."}), 403

        data = request.json
        name = data.get('name', '').strip()
        user_id = data.get('user_id', '').strip()
        password = data.get('password', '')
        email = data.get('email', '').strip() if data.get('email') else None

        if not all([name, user_id, password]):
            return jsonify({"error": "Name, ID, and password are required"}), 400

        # Check if user exists
        if User.query.filter_by(user_id=user_id).first():
            return jsonify({"error": f"Admin ID '{user_id}' already exists"}), 400

        if email and User.query.filter_by(email=email).first():
            return jsonify({"error": f"Email '{email}' already in use"}), 400

        # Create new admin
        pw_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        new_admin = User(
            user_id=user_id,
            name=name,
            email=email,
            role='admin',
            password_hash=pw_hash,
            is_active=True
        )

        db.session.add(new_admin)
        db.session.flush()

        # Log action
        audit = AdminAuditLog(
            action='admin_created',
            admin_id=current_user_id,
            target_id=user_id,
            description=f"Admin created: {name} ({user_id})"
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Admin '{name}' created successfully",
            "admin": {
                "user_id": new_admin.user_id,
                "name": new_admin.name,
                "email": new_admin.email,
                "role": "admin",
                "is_active": True
            }
        }), 201

    except Exception as e:
        print(f"Create Admin Error: {e}")
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# 3C. List All Admins
@app.route('/api/list_admins', methods=['GET'])
def list_admins():
    """Get list of all admin accounts"""
    try:
        admins = User.query.filter_by(role='admin').all()

        admin_list = [{
            "user_id": admin.user_id,
            "name": admin.name,
            "email": admin.email,
            "is_active": admin.is_active,
            "created_at": admin.registration_date.isoformat() + 'Z' if admin.registration_date else None,
            "last_active": admin.last_active.isoformat() + 'Z' if admin.last_active else None
        } for admin in admins]

        return jsonify({"success": True, "admins": admin_list}), 200

    except Exception as e:
        print(f"List Admins Error: {e}")
        return jsonify({"error": str(e)}), 500

# 3G. Deactivate Admin
@app.route('/api/deactivate_admin', methods=['POST'])
def deactivate_admin():
    """Deactivate an admin account"""
    try:
        current_user_id = request.json.get('current_user') or 'ADMIN01'
        data = request.json
        target_user_id = data.get('user_id')

        if not target_user_id:
            return jsonify({"error": "User ID is required"}), 400

        target_admin = User.query.filter_by(user_id=target_user_id, role='admin').first()

        if not target_admin:
            return jsonify({"error": "Admin not found"}), 404

        target_admin.is_active = False

        # Log action
        audit = AdminAuditLog(
            action='admin_deactivated',
            admin_id=current_user_id,
            target_id=target_user_id,
            description=f"Admin deactivated: {target_admin.name}"
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Admin '{target_admin.name}' has been deactivated"
        }), 200

    except Exception as e:
        print(f"Deactivate Admin Error: {e}")
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# 3H. Reactivate Admin
@app.route('/api/reactivate_admin', methods=['POST'])
def reactivate_admin():
    """Reactivate a deactivated admin account"""
    try:
        current_user_id = request.json.get('current_user') or 'ADMIN01'
        data = request.json
        target_user_id = data.get('user_id')

        if not target_user_id:
            return jsonify({"error": "User ID is required"}), 400

        target_admin = User.query.filter_by(user_id=target_user_id, role='admin').first()

        if not target_admin:
            return jsonify({"error": "Admin not found"}), 404

        target_admin.is_active = True

        # Log action
        audit = AdminAuditLog(
            action='admin_activated',
            admin_id=current_user_id,
            target_id=target_user_id,
            description=f"Admin reactivated: {target_admin.name}"
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Admin '{target_admin.name}' has been reactivated"
        }), 200

    except Exception as e:
        print(f"Reactivate Admin Error: {e}")
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# 3H. Delete Admin (Hard Delete)
@app.route('/api/delete_admin', methods=['POST'])
def delete_admin():
    """Permanently delete an admin account"""
    try:
        current_user_id = request.json.get('current_user') or 'ADMIN01'
        admin_user = User.query.filter_by(user_id=current_user_id, role='admin', is_active=True).first()

        if not admin_user:
            return jsonify({"error": "Unauthorized. Only active admins can delete users."}), 403

        target_user_id = request.json.get('user_id', '').strip()

        if not target_user_id:
            return jsonify({"error": "User ID is required"}), 400

        # Prevent self-deletion
        if target_user_id == current_user_id:
            return jsonify({"error": "Cannot delete your own account"}), 400

        target_admin = User.query.filter_by(user_id=target_user_id, role='admin').first()

        if not target_admin:
            return jsonify({"error": "Admin not found"}), 404

        # Store info for audit log before deletion
        target_name = target_admin.name
        target_email = target_admin.email

        # Delete the admin
        db.session.delete(target_admin)

        # Log action
        audit = AdminAuditLog(
            action='admin_deleted',
            admin_id=current_user_id,
            target_id=target_user_id,
            description=f"Admin permanently deleted: {target_name} ({target_email})"
        )
        db.session.add(audit)
        db.session.commit()

        print(f"✅ Admin deleted: {target_name} (ID: {target_user_id})")

        return jsonify({
            "success": True,
            "message": f"Admin '{target_name}' has been permanently deleted"
        }), 200

    except Exception as e:
        print(f"Delete Admin Error: {e}")
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# 3I. Admin Audit Log
@app.route('/api/admin_audit_log', methods=['GET'])
def admin_audit_log():
    """Get admin activity audit log"""
    try:
        logs = AdminAuditLog.query.order_by(AdminAuditLog.created_at.desc()).limit(50).all()

        log_list = [{
            "action": log.action,
            "admin_id": log.admin_id,
            "target_id": log.target_id,
            "description": log.description,
            "created_at": log.created_at.isoformat() + 'Z'
        } for log in logs]

        return jsonify({"success": True, "logs": log_list}), 200

    except Exception as e:
        print(f"Audit Log Error: {e}")
        return jsonify({"error": str(e)}), 500

# 3A. Admin: Get Pending Faculty Registrations
@app.route('/api/admin/pending_faculty_registrations', methods=['GET'])
def get_pending_faculty_registrations():
    """Get all pending faculty registrations for admin approval"""
    try:
        pending_users = User.query.filter_by(role='faculty', registration_status='Pending').order_by(User.registration_submitted_at.desc()).all()

        registrations = []
        for user in pending_users:
            registrations.append({
                "user_id": user.user_id,
                "name": user.name,
                "email": user.email,
                "submitted_at": user.registration_submitted_at.isoformat() + 'Z' if user.registration_submitted_at else None,
                "status": user.registration_status,
                "has_face": user.face_encoding is not None
            })

        return jsonify({
            "success": True,
            "total": len(registrations),
            "registrations": registrations
        }), 200

    except Exception as e:
        print(f"Pending Registrations Error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# 3A2. Admin: Approve Faculty Registration
@app.route('/api/admin/approve_faculty', methods=['POST'])
def approve_faculty_registration():
    """Admin approves a pending faculty registration"""
    data = request.json
    try:
        admin_id = data.get('admin_id', 'ADMIN01')
        user_id = data.get('user_id')

        if not user_id:
            return jsonify({"success": False, "message": "user_id required"}), 400

        user = User.query.filter_by(user_id=user_id, role='faculty').first()

        if not user:
            return jsonify({"success": False, "message": "Faculty not found"}), 404

        if user.registration_status != 'Pending':
            return jsonify({"success": False, "message": f"Registration status is {user.registration_status}, not Pending"}), 400

        # Approve and activate
        user.registration_status = 'Approved'
        user.is_active = True
        user.registration_notes = None
        db.session.commit()

        # Audit log
        audit = AuditLog(
            admin_id=admin_id,
            action=f"Approved faculty registration for {user_id}",
            timestamp=datetime.now(pytz.utc)
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Faculty {user_id} approved and can now login",
            "user_id": user_id
        }), 200

    except Exception as e:
        print(f"Approve Faculty Error: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

# 3A3. Admin: Reject Faculty Registration
@app.route('/api/admin/reject_faculty', methods=['POST'])
def reject_faculty_registration():
    """Admin rejects a pending faculty registration"""
    data = request.json
    try:
        admin_id = data.get('admin_id', 'ADMIN01')
        user_id = data.get('user_id')
        notes = data.get('notes', 'No reason provided').strip()

        if not user_id:
            return jsonify({"success": False, "message": "user_id required"}), 400

        if len(notes) > 300:
            return jsonify({"success": False, "message": "Notes must be 300 characters or less"}), 400

        user = User.query.filter_by(user_id=user_id, role='faculty').first()

        if not user:
            return jsonify({"success": False, "message": "Faculty not found"}), 404

        if user.registration_status != 'Pending':
            return jsonify({"success": False, "message": f"Registration status is {user.registration_status}, not Pending"}), 400

        # Reject and deactivate
        user.registration_status = 'Rejected'
        user.is_active = False
        user.registration_notes = notes
        db.session.commit()

        # Audit log
        audit = AuditLog(
            admin_id=admin_id,
            action=f"Rejected faculty registration for {user_id}. Reason: {notes}",
            timestamp=datetime.now(pytz.utc)
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Faculty {user_id} registration rejected",
            "user_id": user_id
        }), 200

    except Exception as e:
        print(f"Reject Faculty Error: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

# 3B. Register Face - Multi-Angle Capture
@app.route('/api/register_face_multi_angle', methods=['POST'])
def register_face_multi_angle():
    data = request.json
    user_id = data.get('user_id')
    face_images = data.get('face_images')  # Dict: {'front': base64, 'left': base64, ...}

    try:
        # Verify user exists
        user = User.query.filter_by(user_id=user_id).first()
        if not user:
            return jsonify({"success": False, "message": "User not found"}), 404

        if not face_images or len(face_images) < 3:
            return jsonify({"success": False, "message": "Minimum 3 angles required (front, left, right)"}), 400

        # Process all angles and generate averaged encoding
        encodings_list = []
        angles_processed = []

        for angle, base64_image in face_images.items():
            try:
                # Decode base64 image
                image_data = base64_image.split(',')[1] if ',' in base64_image else base64_image
                image_bytes = base64.b64decode(image_data)

                # Get face encoding for this angle
                encoding, location = face_system.get_face_encoding(io.BytesIO(image_bytes))

                if encoding is None:
                    print(f"Failed to process {angle} angle")
                    continue

                encodings_list.append(encoding)
                angles_processed.append(angle)

            except Exception as e:
                print(f"Error processing {angle} image: {e}")
                continue

        # Need at least 3 valid angles for reliable recognition
        if len(encodings_list) < 3:
            return jsonify({
                "success": False,
                "message": f"Only {len(encodings_list)} angles processed successfully. Need at least 3."
            }), 400
        user.face_encoding = pickle.dumps(encodings_list).hex()
        user.face_registered_at = datetime.now(pytz.utc)
        db.session.commit()

        # Audit log (use 'system' fo    r self-registration during first login)
        audit = AuditLog(
            admin_id='SYSTEM',
            action=f"User {user_id} registered face with {len(encodings_list)} angles: {', '.join(angles_processed)}",
            timestamp=datetime.now(pytz.utc)
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Face registered successfully using {len(encodings_list)} angles",
            "angles_used": angles_processed
        })

    except Exception as e:
        print(f"Multi-Angle Face Registration Error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# 3C. Faculty Self-Registration (Login Page)
# 3C. Faculty Self-Registration (Login Page)
@app.route('/api/faculty_self_register', methods=['POST'])
def faculty_self_register():
    """
    Faculty self-registration from login page with face capture.
    Stores registration as PENDING for admin approval.
    """
    data = request.json
    try:
        user_id = data.get('user_id', '').strip().upper()
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')
        face_image = data.get('face_image')  # Single face capture (base64)

        # Validation
        if not all([user_id, name, email, password, face_image]):
            return jsonify({"success": False, "message": "All fields required: user_id, name, email, password, face_image"}), 400

        if len(password) < 6:
            return jsonify({"success": False, "message": "Password must be at least 6 characters"}), 400

        if len(user_id) < 3 or len(user_id) > 20:
            return jsonify({"success": False, "message": "User ID must be 3-20 characters"}), 400

        if len(name) < 2 or len(name) > 100:
            return jsonify({"success": False, "message": "Name must be 2-100 characters"}), 400

        # Check if user already exists
        existing_user = User.query.filter(
            (User.user_id == user_id) | (User.email == email)
        ).first()

        if existing_user:
            if existing_user.user_id == user_id:
                return jsonify({"success": False, "message": "User ID already registered"}), 400
            else:
                return jsonify({"success": False, "message": "Email already registered"}), 400

        # Decode and process face image
        try:
            image_data = face_image.split(',')[1] if ',' in face_image else face_image
            image_bytes = base64.b64decode(image_data)
            encoding, location = face_system.get_face_encoding(image_bytes)

            if encoding is None:
                return jsonify({"success": False, "message": "Face not detected. Ensure good lighting and clear visibility."}), 400
        except Exception as e:
            print(f"Face encoding error in self-register: {e}")
            return jsonify({"success": False, "message": "Face processing failed"}), 400

        # Hash the provided password
        pw_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        # Create new faculty user with PENDING status
        new_user = User(
            user_id=user_id,
            name=name,
            email=email,
            role='faculty',
            face_encoding=pickle.dumps([encoding]).hex(),  # Store as list for consistency
            face_registered_at=datetime.now(pytz.utc),
            registration_status='Pending',  # Faculty starts as PENDING
            registration_submitted_at=datetime.now(pytz.utc),
            is_active=False,  # Can't login until approved
            password_hash=pw_hash
        )

        db.session.add(new_user)
        db.session.commit()

        # Audit log
        audit = AuditLog(
            admin_id='SYSTEM',
            action=f"Faculty {user_id} self-registered. Pending admin approval.",
            timestamp=datetime.now(pytz.utc)
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Registration submitted successfully. Awaiting admin approval.",
            "user_id": user_id,
            "status": "Pending"
        }), 201

    except Exception as e:
        print(f"Faculty Self-Registration Error: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
# 4. Recognize Face (For Login or Attendance)
@app.route('/api/recognize', methods=['POST'])
def recognize():
    data = request.json
    try:
        # Check if face system is loaded
        if not face_system.models_loaded:
            print("❌ ERROR: Face system models not loaded!")
            return jsonify({
                "success": False,
                "error_code": "FACE_SYSTEM_ERROR",
                "message": "Face recognition system is not initialized. Please contact admin."
            }), 503

        if 'image' not in data:
            print("❌ ERROR: No image in request")
            return jsonify({
                "success": False,
                "error_code": "NO_IMAGE",
                "message": "No camera image was received. Please try scanning again."
            }), 400

        # Decode image
        try:
            image_data = base64.b64decode(data['image'].split(',')[1])
            image_bytes = image_data
        except Exception as e:
            print(f"❌ ERROR: Image decoding failed: {e}")
            return jsonify({
                "success": False,
                "error_code": "IMAGE_DECODE_ERROR",
                "message": "Failed to process image. Capture again."
            }), 400

        # Get face encoding and location
        unknown_encoding, face_location = face_system.get_face_encoding(image_bytes)

        if unknown_encoding is None:
             print("❌ Face Detection Failed")
             return jsonify({
                 "success": False,
                 "error_code": "FACE_NOT_DETECTED",
                 "message": "Face not detected clearly. Improve lighting, keep still, and align your face in frame."
             }), 400

        # Load all users from DB
        users = User.query.all()
        known_encodings = {}

        print(f"📊 Loading face encodings for {len(users)} users...")
        for user in users:
            if user.face_encoding:
                try:
                    encoding = user.face_encoding

                    # Decode if stored as hex string (from multi-angle capture)
                    if isinstance(encoding, str):
                        encoding = pickle.loads(bytes.fromhex(encoding))

                    # Normalize to numpy array
# Normalize to numpy array
                    if isinstance(encoding, list):
                        if len(encoding) == 0:
                            continue
                        # Keep the whole list of angles! Don't use np.mean()
                        known_encodings[user.user_id] = encoding
                        print(f"  ✓ Loaded encoding for {user.user_id}")
                    else:
                        known_encodings[user.user_id] = [encoding]
                        print(f"  ✓ Loaded encoding for {user.user_id}")
                except Exception as e:
                    print(f"  ✗ Error processing face encoding for {user.user_id}: {e}")
                    continue

        print(f"📊 Loaded {len(known_encodings)} valid face encodings")
        # Compare
        print(f"🔍 Comparing against {len(known_encodings)} known encodings...")
        best_match_id = face_system.compare_faces(known_encodings, unknown_encoding, threshold=0.45)

        if best_match_id:
            print(f"✓ Face matched to: {best_match_id}")

            # --- FACE MATCH LOCK ---
            # Verify that the recognized face matches the currently logged-in user.
            # The frontend sends 'logged_in_user_id' with every scan request.
            logged_in_user_id = data.get('logged_in_user_id')
            if logged_in_user_id and logged_in_user_id != best_match_id:
                print(f"🚫 FACE MISMATCH: logged_in={logged_in_user_id}, recognized={best_match_id}")
                # Log security alert for identity mismatch
                db.session.add(SecurityAlert(
                    user_id=logged_in_user_id,
                    user_name=f"Logged-in: {logged_in_user_id}",
                    role='faculty',
                    event_code='FACE_IDENTITY_MISMATCH',
                    event_message=f'Face recognized as {best_match_id} but logged in as {logged_in_user_id}. Possible proxy attendance attempt.'
                ))
                db.session.commit()
                return jsonify({
                    "success": False,
                    "error_code": "FACE_MISMATCH",
                    "message": f"Face recognized as a different user ({best_match_id}). You must scan your own face. This incident has been logged."
                }), 403
        else:
            print(f"✗ No face match found (below threshold of 0.45)")

        # --- Location Check (Only for Registered Users) ---
        user_loc = data.get('location')
        if not user_loc or not isinstance(user_loc, dict):
            return jsonify({
                "success": False,
                "error_code": "LOCATION_REQUIRED",
                "message": "Location access is required. Please enable GPS and allow location permission."
            }), 400

        try:
            user_lat = float(user_loc.get('latitude'))
            user_lon = float(user_loc.get('longitude'))

            dist = haversine(user_lat, user_lon, TARGET_LAT, TARGET_LON)
            dist_m = dist * 1000
            print(f"DEBUG: User Loc: ({user_lat}, {user_lon}) | Target: ({TARGET_LAT}, {TARGET_LON}) | Dist: {dist_m:.2f}m")

            # 🔴 FIXED: If user is registered and location enforcement is enabled, enforce boundary check
            # Apply GPS accuracy buffer to prevent false-positive IN BOUNDS
            effective_radius_m = (ALLOWED_RADIUS_KM * 1000) - GPS_ACCURACY_BUFFER_M
            if LOCATION_ENFORCEMENT_ENABLED and best_match_id and dist_m > effective_radius_m:
                 return jsonify({
                     "success": False,
                     "error_code": "OUT_OF_BOUNDS",
                     "message": f"Out of campus boundary. You are {dist_m:.0f}m away. Move inside {effective_radius_m:.0f}m and scan again."
                 }), 403

        except (ValueError, TypeError):
             return jsonify({
                 "success": False,
                 "error_code": "LOCATION_INVALID",
                 "message": "Invalid location coordinates received. Re-enable location and try again."
             }), 400

        if best_match_id:
            user = User.query.filter_by(user_id=best_match_id).first()

            # Check if faculty is approved and active
            if user.role == 'faculty':
                if user.registration_status != 'Approved':
                    return jsonify({
                        "success": False,
                        "error_code": "INVALID_USER",
                        "message": f"Your registration status is {user.registration_status}. Only approved faculty can mark attendance."
                    }), 403
                if not user.is_active:
                    return jsonify({
                        "success": False,
                        "error_code": "INVALID_USER",
                        "message": "Your account is deactivated. Contact admin."
                    }), 403

            # Auto-Mark Attendance if recognized
            # Use UTC for storage, Local for logic
            now_utc = datetime.now(pytz.utc)
            # Assuming 'Asia/Kolkata' or similar based on name 'Krishna', but let's default to a config or UTC+5:30 for now or just generic local.
            # Ideally, use pytz.timezone('Asia/Kolkata').localize(datetime.now())
            # For this 'deployment ready' request, I'll use a fixed standard offset or server local if simpler, but user asked for Timezones.
            # FIX: Convert UTC to Local for Display
            local_tz = pytz.timezone('Asia/Kolkata') # Example, or use 'UTC'
            now_local = now_utc.astimezone(local_tz)

            today_str = now_local.strftime('%Y-%m-%d')
            time_str = now_local.strftime('%H:%M:%S')

            # --- 4. Holiday Check ---
            holiday = Holiday.query.filter_by(date=today_str).first()
            if holiday:
                 return jsonify({
                    "success": True,
                    "user": {
                        "id": user.user_id,
                        "user_id": user.user_id,
                        "name": user.name,
                        "role": user.role
                    },
                    "location": face_location,
                    "attendance": { "status": f"Holiday: {holiday.name}", "type": "HOLIDAY" }
                })

            # Check if already checked in today (Debounce logic)
            today_log = AttendanceLog.query.filter_by(user_id=user.user_id, date=today_str).first()

            log_type = 'IN'
            msg = "Marked Present"

            if today_log:
                if today_log.time_out:
                    msg = "Attendance Complete"
                    log_type = "already_done"

                else:
                    # Has checked IN, now checking OUT (if cooldown passes)
                    min_gap = 10

                    time_diff = (now_utc.replace(tzinfo=None) - today_log.timestamp_in).total_seconds()

                    if time_diff < min_gap:
                         msg = f"Wait {int(min_gap-time_diff)}s to Checkout"
                         log_type = "cooldown"
                    else:
                        second = classify_second_mark(today_log.check_in_status or today_log.status, now_local)
                        if not second["allowed"]:
                            log_type = "ignored"
                            msg = second["message"]
                        else:
                            log_type = 'OUT'
                            today_log.time_out = time_str
                            today_log.timestamp_out = now_utc
                            today_log.check_out_status = second["out_status"]
                            today_log.check_out_period = second["period"]
                            today_log.status = second["final_status"]
                            db.session.commit()
                            msg = second["message"]
            else:
                first = classify_first_mark(now_local)
                if not first["allowed"]:
                    return jsonify({
                        "success": True,
                        "user": {
                            "id": user.user_id,
                            "user_id": user.user_id,
                            "name": user.name,
                            "role": user.role
                        },
                        "location": face_location,
                        "attendance": {
                            "status": first["message"],
                            "type": "IGNORED"
                        }
                    })

                # Check for approved permissions
                has_permission = has_approved_permission(user.user_id, today_str)

                new_log = AttendanceLog(
                    user_id=user.user_id,
                    date=today_str,
                    time_in=time_str,
                    time_out=None,
                    check_in_status=first["status"],
                    check_in_period=first["period"],
                    check_out_status=None,
                    check_out_period=None,
                    timestamp_in=now_utc,
                    status=first["status"],
                    late_permission_approved=has_permission
                )
                db.session.add(new_log)
                db.session.commit()
                msg = first["message"]
                log_type = 'IN'

            return jsonify({
                "success": True,
                "user": {
                    "id": user.user_id,
                    "user_id": user.user_id,
                    "name": user.name,
                    "role": user.role
                },
                "location": face_location,
                "attendance": {
                    "status": msg,
                    "type": log_type
                }
            })
        else:
            db.session.add(SecurityAlert(
                user_id=None,
                user_name='Unknown Face',
                role='unknown',
                event_code='INVALID_USER_SCAN',
                event_message='Unregistered or invalid user attempted face scan.'
            ))
            db.session.commit()
            return jsonify({
                "success": False,
                "error_code": "INVALID_USER",
                "message": "Invalid user or unregistered face. Please contact admin if you are a valid staff member."
            }), 401

    except Exception as e:
        print(f"Recognition Error: {e}")
        return jsonify({
            "success": False,
            "error_code": "SERVER_ERROR",
            "message": "Server error while validating attendance. Please retry in a moment.",
            "error": str(e)
        }), 500


@app.route('/api/location_heartbeat', methods=['POST'])
def location_heartbeat():
    """
    Receives periodic location pings from logged-in mobile users.
    Enforces lunch-window boundary policy and returns alert hints for frontend.
    """
    data = request.json or {}
    user_id = data.get('user_id')
    user_loc = data.get('location')
    device_status = data.get('device_status') or {}

    if not user_id:
        return jsonify({"success": False, "message": "user_id is required"}), 400

    user = User.query.filter_by(user_id=user_id).first()
    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404

    # FIXED: WebView's navigator.onLine is UNRELIABLE (returns false when app is in background).
    # For FACULTY users, the native Android service is the ONLY source of truth for network/location.
    # We ignore network_on and location_on from WebView heartbeats entirely for faculty.
    # Admin users don't have native tracking, so we just skip device_status claims altogether.

    # NOTE: We deliberately do NOT set NETWORK_OFF or LOCATION_OFF from WebView heartbeats.
    # The native service (/api/faculty/location) handles that with real Android system data.

    # Handle explicit logout signal — remove user from LivePresence entirely
    if data.get('logout'):
        presence = LivePresence.query.filter_by(user_id=user_id).first()
        if presence:
            db.session.delete(presence)
            db.session.commit()
            print(f"[Logout] Removed LivePresence for {user_id}")
        return jsonify({"success": True, "message": "Logged out"})

    if not user_loc or not isinstance(user_loc, dict):
        return jsonify({"success": False, "message": "Valid location payload is required"}), 400

    try:
        user_lat = float(user_loc.get('latitude'))
        user_lon = float(user_loc.get('longitude'))
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Invalid coordinates"}), 400

    dist_km = haversine(user_lat, user_lon, TARGET_LAT, TARGET_LON)
    dist_m = dist_km * 1000
    
    # 🔴 FIXED: Apply GPS accuracy buffer for stricter boundary
    # GPS can be off by 5-20m, so we reduce effective boundary by GPS_ACCURACY_BUFFER_M
    effective_radius_m = (ALLOWED_RADIUS_KM * 1000) - GPS_ACCURACY_BUFFER_M
    in_bounds = dist_m <= effective_radius_m

    # If location enforcement is disabled, skip boundary checks
    if not LOCATION_ENFORCEMENT_ENABLED:
        in_bounds = True

    print(f"DEBUG HEARTBEAT: {user.user_id} | Loc: ({user_lat:.6f}, {user_lon:.6f}) | Target: ({TARGET_LAT:.6f}, {TARGET_LON:.6f}) | Dist: {dist_m:.2f}m | Nominal Limit: {ALLOWED_RADIUS_KM*1000:.0f}m | Effective Limit: {effective_radius_m:.0f}m (GPS buffer: {GPS_ACCURACY_BUFFER_M}m) | InBounds: {in_bounds} | Enforcement: {LOCATION_ENFORCEMENT_ENABLED}")

    now_utc = datetime.now(pytz.utc)
    local_tz = pytz.timezone('Asia/Kolkata')
    now_local = now_utc.astimezone(local_tz)
    today_str = now_local.strftime('%Y-%m-%d')
    flags = get_time_policy_flags(now_local)

    alert_message = None
    alert_code = None
    marked_absent = False

    # 12:50 PM pre-alert for lunch start policy reminder.
    if flags["is_lunch_pre_alert"]:
        alert_message = "Lunch window starts at 01:00 PM and ends at 01:40 PM. Return before 01:40 PM."
        alert_code = "LUNCH_START_REMINDER"

    # 10-minute reminder before lunch close (01:30 PM to 01:40 PM).
    if flags["is_return_pre_alert"]:
        alert_message = "10-minute reminder: Lunch free-exit ends at 01:40 PM. Please return inside campus now."
        alert_code = "LUNCH_END_REMINDER"

    # Lunch free-exit window: ALLOW free exit but DON'T force IN status
    # If you're physically OUT during lunch, you stay MARKED AS OUT
    # (You're allowed to be out, but not marked as "present in bounds")
    if flags["is_lunch_window"] and not in_bounds:
        # Don't modify in_bounds - keep actual location
        print(f"DEBUG HEARTBEAT: {user.user_id} is outside during lunch (allowed), but STILL MARKED AS OUT for accuracy.")

    # After 1:40 PM, outside campus is marked absent until check-out window starts.
    if flags["is_post_lunch_enforcement"] and not in_bounds:
        today_log = AttendanceLog.query.filter_by(user_id=user.user_id, date=today_str).first()
        if today_log and not today_log.time_out:
            today_log.status = "Absent"
            db.session.commit()
            marked_absent = True
            alert_message = "Outside campus after 01:40 PM. Marked Absent. Contact admin if extra time is required."
            alert_code = "POST_LUNCH_OUT_OF_BOUNDS_ABSENT"
        elif not today_log:
            alert_message = "Outside campus after 01:40 PM. Check-in and stay within campus boundary."
            alert_code = "POST_LUNCH_OUT_OF_BOUNDS"

    if in_bounds:
        presence_code = 'OK'
        presence_msg = 'Inside campus boundary.'
    else:
        presence_code = 'OUT_OF_BOUNDS'
        presence_msg = f'Outside campus boundary ({round(dist_km * 1000, 2)} m from center).'

    upsert_live_presence(
        user,
        status_code=presence_code,
        status_message=presence_msg,
        source='heartbeat',
        latitude=user_lat,
        longitude=user_lon,
        distance_m=round(dist_km * 1000, 2),
        in_bounds=in_bounds
    )
    db.session.commit()

    return jsonify({
        "success": True,
        "server_time": now_local.strftime('%H:%M:%S'),
        "date": today_str,
        "distance_m": round(dist_km * 1000, 2),
        "in_bounds": in_bounds,
        "marked_absent": marked_absent,
        "time_flags": flags,
        "alert": alert_message,
        "alert_code": alert_code
    })
    
@app.route('/api/faculty/location', methods=['POST'])
def faculty_location():
    try:
        data = request.json or {}
        user_id = data.get('user_id')

        if not user_id:
            return jsonify({"success": False, "message": "user_id required"}), 400

        user = User.query.filter_by(user_id=user_id).first()
        if not user or user.role != 'faculty':
            return jsonify({"success": False, "message": "Unauthorized"}), 403

        # --- 1. EXTRACT NATIVE TRUTH FLAGS ---
        location_on = data.get('location_on', True)
        
        # FIXED: If this endpoint is successfully hit, the device HAS network access.
        # Android's ConnectivityManager can be flaky in Doze mode or when GPS toggles.
        # We ignore the payload's network_status claim and definitively mark it online.
        network_status = 'online'
        
        try:
            user_lat = float(data.get('latitude', 0.0))
            user_lon = float(data.get('longitude', 0.0))
            accuracy = float(data.get('accuracy', 0.0))
        except:
            user_lat, user_lon, accuracy = 0.0, 0.0, 0.0

        # --- 2. DETERMINE TRUE STATUS ---
        if not location_on:
            status_code = 'LOCATION_OFF'
            status_message = 'GPS disabled on device'
            in_bounds = False
            user_lat = None  
            user_lon = None
            dist_km = None
            
        elif user_lat == 0.0 and user_lon == 0.0:
            status_code = 'ACQUIRING_GPS'
            status_message = 'Searching for GPS signal...'
            in_bounds = False
            user_lat = None
            user_lon = None
            dist_km = None
            
        else:
            # Normal GPS tracking
            dist_km = haversine(user_lat, user_lon, TARGET_LAT, TARGET_LON)
            # 🔴 FIXED: Apply GPS accuracy buffer for stricter boundary (line 2840)
            effective_radius_m = (ALLOWED_RADIUS_KM * 1000) - GPS_ACCURACY_BUFFER_M
            in_bounds = (dist_km * 1000 <= effective_radius_m) if LOCATION_ENFORCEMENT_ENABLED else True
            status_code = 'OK' if in_bounds else 'OUT_OF_BOUNDS'
            status_message = 'In campus' if in_bounds else f'Outside boundary ({round(dist_km * 1000, 2)}m away)'

        # --- 3. EVENT-DRIVEN STATE LOGGING ---
        # Track 3 independent dimensions: NETWORK, LOCATION (GPS toggle), BOUNDS
        # Only insert a DeviceStateLog row when a specific dimension changes.
        now_utc = datetime.now(pytz.utc)
        local_tz = pytz.timezone('Asia/Kolkata')
        now_local = now_utc.astimezone(local_tz)
        today_str = now_local.strftime('%Y-%m-%d')

        current_network = network_status                        # 'online' or 'offline'
        current_location = 'active' if location_on else 'inactive'   # GPS toggle
        current_bounds = 'in' if in_bounds else 'out'           # geofence

        # Common coordinates for the log entry
        log_lat = user_lat
        log_lon = user_lon
        log_dist = round(dist_km * 1000, 2) if dist_km is not None else None

        state_logged = False

        # Helper: get last DeviceStateLog for a specific event_type
        def get_last_device_state(evt_type):
            return DeviceStateLog.query.filter_by(
                user_id=user_id, date=today_str, event_type=evt_type
            ).order_by(DeviceStateLog.timestamp.desc()).first()

        # --- DIMENSION 1: NETWORK (online ↔ offline) ---
        last_net = get_last_device_state('NETWORK')
        prev_net = last_net.new_state if last_net else 'online'
        if prev_net != current_network:
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='NETWORK',
                old_state=prev_net, new_state=current_network,
                latitude=log_lat, longitude=log_lon, distance_m=log_dist,
                accuracy_m=accuracy, timestamp=now_utc
            ))
            state_logged = True
            print(f"[StateLog] {user_id} NETWORK: {prev_net} → {current_network}")

        # --- DIMENSION 2: LOCATION / GPS toggle (active ↔ inactive) ---
        last_loc = get_last_device_state('LOCATION')
        prev_loc = last_loc.new_state if last_loc else 'active'
        if prev_loc != current_location:
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='LOCATION',
                old_state=prev_loc, new_state=current_location,
                latitude=log_lat, longitude=log_lon, distance_m=log_dist,
                accuracy_m=accuracy, timestamp=now_utc
            ))
            state_logged = True
            print(f"[StateLog] {user_id} LOCATION: {prev_loc} → {current_location}")

        # --- DIMENSION 3: BOUNDS / geofence (in ↔ out) ---
        last_bnd = get_last_device_state('BOUNDS')
        prev_bnd = last_bnd.new_state if last_bnd else 'in'
        if prev_bnd != current_bounds:
            db.session.add(DeviceStateLog(
                user_id=user_id, date=today_str, event_type='BOUNDS',
                old_state=prev_bnd, new_state=current_bounds,
                latitude=log_lat, longitude=log_lon, distance_m=log_dist,
                accuracy_m=accuracy, timestamp=now_utc
            ))
            state_logged = True
            print(f"[StateLog] {user_id} BOUNDS: {prev_bnd} → {current_bounds}")

        # Also write to the legacy LocationLog on any state change (for backward compat)
        if state_logged and log_lat is not None and log_lon is not None:
            db.session.add(LocationLog(
                user_id=user_id, date=today_str,
                latitude=log_lat, longitude=log_lon,
                distance_m=log_dist if log_dist is not None else 0.0,
                in_bounds=in_bounds, network_status=current_network,
                accuracy_m=accuracy, timestamp=now_utc
            ))

        # --- 4. SAVE TO LivePresence (with NATIVE TRUTH fields) ---
        upsert_live_presence(
            user=user,
            status_code=status_code,
            status_message=status_message,
            source='native_tracker',
            latitude=user_lat,
            longitude=user_lon,
            distance_m=round(dist_km * 1000, 2) if dist_km is not None else None,
            in_bounds=in_bounds,
            native_network_status=network_status,
            native_location_enabled=bool(location_on)
        )
        db.session.commit()

        # 🟡 PIGGYBACK: Run inactive device wakeup in background
        # This runs AFTER response is sent, costs nothing, no scheduler needed
        Thread(target=ping_inactive_devices, daemon=True).start()

        return jsonify({"success": True, "state_logged": state_logged}), 200

    except Exception as e:
        print(f"[LocationPing ERROR] {str(e)}")
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/force_offline', methods=['POST'])
def force_offline():
    """Logout kill-switch: Force a user's LivePresence to Inactive/Offline.
    Called by:
    1. Kotlin LocationTrackingService on ACTION_STOP (fireForceOffline)
    2. JS appLogout() as belt-and-suspenders
    """
    try:
        data = request.json or {}
        user_id = (data.get('user_id') or '').strip()

        if not user_id:
            return jsonify({"success": False, "message": "user_id required"}), 400

        # 🔴 LOG FINAL DEVICE STATES: Transition to offline/inactive/out
        log_final_device_states(user_id)

        # Clear LivePresence record
        presence = LivePresence.query.filter_by(user_id=user_id).first()
        if presence:
            db.session.delete(presence)
            db.session.commit()
            print(f"🔴 Force-offline: {user_id} marked as OFFLINE and LivePresence cleared")
        else:
            print(f"🔴 Force-offline: No LivePresence found for {user_id} (already cleaned up)")

        return jsonify({"success": True, "message": f"User {user_id} forced offline"}), 200

    except Exception as e:
        db.session.rollback()
        print(f"[ForceOffline ERROR] {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500
        
@app.route('/api/admin/live_locations', methods=['GET'])
def admin_live_locations():
    cleanup_stale_presence()
    admin_id = request.args.get('admin_id')
    if not admin_id:
        return jsonify({"success": False, "message": "admin_id is required"}), 400

    admin = User.query.filter_by(user_id=admin_id).first()
    if not admin or admin.role != 'admin':
        return jsonify({"success": False, "message": "Unauthorized"}), 403

    now_utc = datetime.utcnow()
    # FIXED: Use 5 minutes instead of 90 seconds.
    # The native service sends every 10s, but Android Doze/background can delay heartbeats
    # significantly. 5 minutes prevents false offline during normal phone use (home button,
    # screen lock, app switching). Real outages will still be detected within 5 minutes.
    stale_cutoff = now_utc - timedelta(seconds=300)

    presences = LivePresence.query.filter(
        LivePresence.role.in_(['faculty', 'admin']),
        LivePresence.user_id != 'ADMIN01'
    ).all()
    map_points = []
    fault_alerts = []

    # Get current time for policy calculation
    now_utc = datetime.utcnow()
    local_tz = pytz.timezone('Asia/Kolkata')
    now_local = now_utc.astimezone(local_tz)
    flags = get_time_policy_flags(now_local)

    for p in presences:
        effective_code = p.status_code or 'UNKNOWN'
        effective_message = p.status_message or 'Status unavailable.'

        # --- If user explicitly logged out (force_offline sets status_code='OFFLINE'),
        # always show them as offline regardless of timing ---
        if effective_code == 'OFFLINE':
            is_heartbeat_fresh = False
            is_network_on = False
            is_location_on = False
        else:
            # --- INDEPENDENTLY determine network and location status ---
            # Network ON = heartbeat arrived within the last 5 minutes
            is_heartbeat_fresh = bool(p.last_seen and p.last_seen >= stale_cutoff)
            is_network_on = is_heartbeat_fresh

            # Location ON = native GPS flag says so AND heartbeat is fresh
            native_loc = getattr(p, 'location_enabled', True)
            if native_loc is None:
                native_loc = True
            is_location_on = bool(native_loc) and is_heartbeat_fresh

        # Build fault list (multiple faults can coexist)
        faults = []
        if not is_heartbeat_fresh:
            faults.append(('STALE', 'No heartbeat received recently. Device may be offline.'))
        if effective_code != 'OFFLINE':
            native_loc = getattr(p, 'location_enabled', True)
            if native_loc is None:
                native_loc = True
            if not native_loc and is_heartbeat_fresh:
                faults.append(('LOCATION_OFF', 'GPS disabled on device (native report).'))
            if effective_code == 'OUT_OF_BOUNDS' or (effective_code == 'OK' and not p.in_bounds):
                faults.append(('OUT_OF_BOUNDS', effective_message))

        # Pick the most critical fault for display
        if effective_code == 'OFFLINE':
            effective_code = 'OFFLINE'
            effective_message = 'User logged out.'
        elif not is_heartbeat_fresh:
            effective_code = 'STALE'
            effective_message = 'No heartbeat received recently. Device may be offline.'
        elif not getattr(p, 'location_enabled', True):
            effective_code = 'LOCATION_OFF'
            effective_message = 'GPS disabled on device (native report).'
        elif not p.in_bounds:
            effective_code = 'OUT_OF_BOUNDS'
            effective_message = p.status_message or 'Outside campus boundary.'
        # else: keep the stored effective_code (OK, ACQUIRING_GPS, etc.)

        # Calculate policy status based on time windows
        policy_in_bounds = p.in_bounds
        if not p.in_bounds:
            if flags["is_lunch_window"]:
                policy_in_bounds = True
            elif flags["is_post_lunch_enforcement"]:
                policy_in_bounds = False

        final_status = effective_code

        # Include ALL active tracking users in map_points
        map_points.append({
            "user_id": p.user_id,
            "name": p.name,
            "role": p.role,
            "latitude": p.latitude,
            "longitude": p.longitude,
            "distance_m": p.distance_m,
            "last_seen": p.last_seen.isoformat() + 'Z' if p.last_seen else None,
            "status": final_status,
            "in_bounds": p.in_bounds,
            "policy_in_bounds": policy_in_bounds,
            "device_status": {
                "network_on": is_network_on,
                "location_on": is_location_on
            }
        })

        # Emit a fault alert for EACH fault dimension (not just the primary one)
        for fault_code, fault_msg in faults:
            fault_alerts.append({
                "fault_key": f"{p.user_id}:{fault_code}",
                "user_id": p.user_id,
                "name": p.name,
                "role": p.role,
                "code": fault_code,
                "message": fault_msg,
                "last_seen": p.last_seen.isoformat() + 'Z' if p.last_seen else None
            })
            print(f"DEBUG: Fault Alert Added: {p.user_id} | {fault_code} | {fault_msg}")

    recent_security_alerts = SecurityAlert.query.filter(
        SecurityAlert.created_at >= (now_utc - timedelta(minutes=30)),
        SecurityAlert.user_id != 'ADMIN01'
    ).order_by(SecurityAlert.created_at.desc()).limit(100).all()

    event_alerts = [
        {
            "fault_key": f"EVENT:{a.id}",
            "user_id": a.user_id,
            "name": a.user_name,
            "role": a.role,
            "code": a.event_code,
            "message": a.event_message,
            "created_at": a.created_at.isoformat() + 'Z' if a.created_at else None
        }
        for a in recent_security_alerts
    ]

    # NEW: Get inactive users (logged out / not in LivePresence)
    all_users = User.query.filter(
        User.role.in_(['faculty', 'admin']),
        User.user_id != 'ADMIN01'
    ).all()

    active_user_ids = {p.user_id for p in presences}
    inactive_users = []

    for user in all_users:
        if user.user_id not in active_user_ids:
            # Get last activity timestamp for inactive user from AttendanceLog
            last_log = AttendanceLog.query.filter_by(user_id=user.user_id).order_by(AttendanceLog.timestamp_out.desc(), AttendanceLog.timestamp_in.desc()).first()
            last_activity = None
            if last_log:
                # Use check-out time if available, else check-in time
                last_activity = (last_log.timestamp_out or last_log.timestamp_in).isoformat() + 'Z' if (last_log.timestamp_out or last_log.timestamp_in) else None

            inactive_users.append({
                "user_id": user.user_id,
                "name": user.name,
                "role": user.role,
                "latitude": None,
                "longitude": None,
                "distance_m": None,
                "last_seen": last_activity,  # NEW: Get from AttendanceLog
                "status": "OFFLINE",
                "in_bounds": False,
                "policy_in_bounds": False,
                "device_status": {
                    "network_on": False,
                    "location_on": False
                }
            })

    return jsonify({
        "success": True,
        "target": {
            "latitude": TARGET_LAT,
            "longitude": TARGET_LON,
            "radius_m": round(ALLOWED_RADIUS_KM * 1000, 2)
        },
        "map_points": map_points,
        "inactive_users": inactive_users,  # NEW: Include users who are not currently tracking
        "fault_alerts": fault_alerts,
        "event_alerts": event_alerts,
        "server_time": now_utc.isoformat() + 'Z'
    })


@app.route('/api/admin/permissions', methods=['POST'])
def admin_create_permission():
    """Admin creates permission record after offline faculty request."""
    try:
        data = request.json or {}
        admin_id = (data.get('admin_id') or '').strip()
        admin_user = require_active_admin(admin_id)
        if not admin_user:
            return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403

        user_id = (data.get('user_id') or '').strip()
        perm_type = (data.get('type') or '').strip().upper()
        date_str = (data.get('date') or datetime.now().strftime('%Y-%m-%d')).strip()
        start_time = normalize_time_hhmm(data.get('start_time'))
        end_time = normalize_time_hhmm(data.get('end_time'))
        is_full_day = bool(data.get('is_full_day', False))
        custom_days = normalize_custom_days(data.get('custom_days'))
        reason = (data.get('reason') or '').strip()
        status = (data.get('status') or 'Pending').strip().title()

        if not user_id or perm_type not in {'LP', 'EP'}:
            return jsonify({"success": False, "message": "Valid user_id and type (LP/EP) are required."}), 400
        if status not in {'Pending', 'Approved', 'Rejected'}:
            return jsonify({"success": False, "message": "status must be Pending, Approved, or Rejected."}), 400

        try:
            datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({"success": False, "message": "date must be in YYYY-MM-DD format."}), 400

        if not is_full_day:
            if not start_time or not end_time:
                return jsonify({"success": False, "message": "start_time and end_time are required unless full-day permission is selected."}), 400
            start_obj = datetime.strptime(start_time, '%H:%M')
            end_obj = datetime.strptime(end_time, '%H:%M')
            if end_obj <= start_obj:
                return jsonify({"success": False, "message": "end_time must be later than start_time."}), 400
        else:
            start_time = None
            end_time = None

        faculty = User.query.filter_by(user_id=user_id).first()
        if not faculty or faculty.role not in ['faculty', 'student', 'admin']:
            return jsonify({"success": False, "message": "Target user not found."}), 404

        permission = PermissionRequest(
            user_id=user_id,
            type=perm_type,
            date=date_str,
            start_time=start_time,
            end_time=end_time,
            is_full_day=is_full_day,
            custom_days=custom_days,
            reason=reason,
            status=status
        )
        db.session.add(permission)

        db.session.add(AdminAuditLog(
            action='permission_created',
            admin_id=admin_id,
            target_id=user_id,
            description=f"Permission {perm_type} created for {user_id} on {date_str} ({'Full Day' if is_full_day else f'{start_time}-{end_time}'}) with status {status}."
        ))
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Permission request recorded by admin.",
            "permission": {
                "id": permission.id,
                "user_id": permission.user_id,
                "type": permission.type,
                "date": permission.date,
                "start_time": permission.start_time,
                "end_time": permission.end_time,
                "is_full_day": bool(permission.is_full_day),
                "custom_days": permission.custom_days,
                "reason": permission.reason,
                "status": permission.status
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/admin/permissions', methods=['GET'])
def admin_list_permissions():
    """Admin views permission requests and decisions."""
    try:
        admin_id = (request.args.get('admin_id') or '').strip()
        admin_user = require_active_admin(admin_id)
        if not admin_user:
            return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403

        status_filter = (request.args.get('status') or '').strip().title()
        user_filter = (request.args.get('user_id') or '').strip()
        date_filter = (request.args.get('date') or '').strip()

        query = PermissionRequest.query
        if status_filter in {'Pending', 'Approved', 'Rejected'}:
            query = query.filter(PermissionRequest.status == status_filter)
        if user_filter:
            query = query.filter(PermissionRequest.user_id == user_filter)
        if date_filter:
            query = query.filter(PermissionRequest.date == date_filter)

        permissions = query.order_by(PermissionRequest.id.desc()).all()
        user_ids = list({p.user_id for p in permissions})
        users = User.query.filter(User.user_id.in_(user_ids)).all() if user_ids else []
        user_map = {u.user_id: u for u in users}

        return jsonify({
            "success": True,
            "permissions": [
                {
                    "id": p.id,
                    "user_id": p.user_id,
                    "name": user_map[p.user_id].name if p.user_id in user_map else p.user_id,
                    "role": user_map[p.user_id].role if p.user_id in user_map else '-',
                    "type": p.type,
                    "date": p.date,
                    "start_time": p.start_time,
                    "end_time": p.end_time,
                    "is_full_day": bool(p.is_full_day),
                    "custom_days": p.custom_days,
                    "reason": p.reason,
                    "status": p.status
                }
                for p in permissions
            ]
        }), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/admin/permissions/<int:permission_id>/decision', methods=['POST'])
def admin_decide_permission(permission_id):
    """Admin approves/rejects permission after in-person faculty request."""
    try:
        data = request.json or {}
        admin_id = (data.get('admin_id') or '').strip()
        admin_user = require_active_admin(admin_id)
        if not admin_user:
            return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403

        decision = (data.get('decision') or '').strip().title()
        decision_reason = (data.get('decision_reason') or '').strip()
        if decision not in {'Approved', 'Rejected'}:
            return jsonify({"success": False, "message": "decision must be Approved or Rejected."}), 400

        permission = PermissionRequest.query.get(permission_id)
        if not permission:
            return jsonify({"success": False, "message": "Permission request not found."}), 404

        permission.status = decision
        if decision_reason:
            existing_reason = permission.reason or ''
            permission.reason = (existing_reason + f" | Admin note: {decision_reason}").strip(' |')

        db.session.add(AdminAuditLog(
            action='permission_decision',
            admin_id=admin_id,
            target_id=permission.user_id,
            description=f"Permission #{permission.id} ({permission.type}) marked {decision}."
        ))
        create_permission_decision_chat_message(permission, admin_id, decision, decision_reason or None)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Permission request {decision.lower()}.",
            "permission": {
                "id": permission.id,
                "user_id": permission.user_id,
                "type": permission.type,
                "date": permission.date,
                "start_time": permission.start_time,
                "end_time": permission.end_time,
                "is_full_day": bool(permission.is_full_day),
                "custom_days": permission.custom_days,
                "reason": permission.reason,
                "status": permission.status
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

# 4B. Delete Permission Request (Admin Only)
@app.route('/api/admin/permissions/<int:permission_id>', methods=['DELETE'])
def delete_permission(permission_id):
    """Admin deletes a permission request (pending or decided)."""
    try:
        data = request.json or {}
        admin_id = (data.get('admin_id') or '').strip()
        admin_user = require_active_admin(admin_id)
        if not admin_user:
            return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403

        permission = PermissionRequest.query.get(permission_id)
        if not permission:
            return jsonify({"success": False, "message": "Permission request not found."}), 404

        perm_type = permission.type
        perm_user_id = permission.user_id

        db.session.add(AdminAuditLog(
            action='permission_deleted',
            admin_id=admin_id,
            target_id=permission.user_id,
            description=f"Permission #{permission.id} ({permission.type}) on {permission.date} deleted."
        ))

        db.session.delete(permission)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Permission request deleted successfully."
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


# Chat-Integrated Permission Endpoints
@app.route('/api/permissions/user/<user_id>', methods=['GET'])
def get_user_permissions(user_id):
    """Get all permission requests for a specific user."""
    try:
        permissions = PermissionRequest.query.filter_by(user_id=user_id).order_by(PermissionRequest.created_at.desc()).all()
        
        return jsonify({
            "success": True,
            "permissions": [
                {
                    "id": p.id,
                    "user_id": p.user_id,
                    "type": p.type,
                    "custom_type": p.custom_type,
                    "date": p.date,
                    "start_time": p.start_time,
                    "end_time": p.end_time,
                    "is_full_day": p.is_full_day,
                    "reason": p.reason,
                    "document_path": p.document_path,
                    "status": p.status,
                    "admin_notes": p.admin_notes,
                    "approved_by": p.approved_by,
                    "created_at": p.created_at.isoformat() if p.created_at else None,
                    "updated_at": p.updated_at.isoformat() if p.updated_at else None
                }
                for p in permissions
            ]
        }), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/permissions/<int:permission_id>/approve', methods=['POST'])
def approve_permission(permission_id):
    """Admin approves a permission request."""
    try:
        data = request.json or {}
        admin_id = (data.get('admin_id') or '').strip()
        admin_notes = (data.get('admin_notes') or '').strip()

        # Verify admin
        admin_user = User.query.filter_by(user_id=admin_id, role='admin', is_active=True).first()
        if not admin_user:
            return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403

        permission = PermissionRequest.query.get(permission_id)
        if not permission:
            return jsonify({"success": False, "message": "Permission request not found."}), 404

        permission.status = 'Approved'
        permission.approved_by = admin_id
        permission.admin_notes = admin_notes
        permission.updated_at = datetime.utcnow()

        db.session.add(permission)
        create_permission_decision_chat_message(permission, admin_id, 'Approved', admin_notes or None)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Permission approved successfully!",
            "permission": {
                "id": permission.id,
                "status": permission.status,
                "approved_by": permission.approved_by,
                "admin_notes": permission.admin_notes,
                "updated_at": permission.updated_at.isoformat()
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/permissions/<int:permission_id>/reject', methods=['POST'])
def reject_permission(permission_id):
    """Admin rejects a permission request."""
    try:
        data = request.json or {}
        admin_id = (data.get('admin_id') or '').strip()
        rejection_reason = (data.get('rejection_reason') or '').strip()

        # Verify admin
        admin_user = User.query.filter_by(user_id=admin_id, role='admin', is_active=True).first()
        if not admin_user:
            return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403

        permission = PermissionRequest.query.get(permission_id)
        if not permission:
            return jsonify({"success": False, "message": "Permission request not found."}), 404

        permission.status = 'Rejected'
        permission.approved_by = admin_id
        if rejection_reason:
            permission.admin_notes = rejection_reason
        permission.updated_at = datetime.utcnow()

        db.session.add(permission)
        create_permission_decision_chat_message(permission, admin_id, 'Rejected', rejection_reason or None)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Permission rejected.",
            "permission": {
                "id": permission.id,
                "status": permission.status,
                "approved_by": permission.approved_by,
                "admin_notes": permission.admin_notes,
                "updated_at": permission.updated_at.isoformat()
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

# 5. Manual Attendance (Admin Only)
@app.route('/api/mark_attendance', methods=['POST'])
def mark_attendance():
    data = request.json or {}
    user_id = data.get('user_id')
    admin_id = (data.get('admin_id') or '').strip()

    admin_user = require_active_admin(admin_id)
    if not admin_user:
        return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403

    user = User.query.filter_by(user_id=user_id).first()

    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404

    # Use UTC -> Local
    now_utc = datetime.now(pytz.utc)
    local_tz = pytz.timezone('Asia/Kolkata')
    now_local = now_utc.astimezone(local_tz)

    today_str = now_local.strftime('%Y-%m-%d')
    time_str = now_local.strftime('%H:%M:%S')

    # Check if already checked in today
    today_log = AttendanceLog.query.filter_by(user_id=user_id, date=today_str).first()

    msg = ""
    log_type = "IN"

    if today_log:
        if today_log.time_out:
            return jsonify({"success": False, "message": "User already completed attendance for today"}), 400
        else:
             # Check out using configured second-mark rules
             second = classify_second_mark(today_log.check_in_status or today_log.status, now_local)
             if not second["allowed"]:
                 return jsonify({"success": False, "message": second["message"]}), 400

             log_type = 'OUT'
             today_log.time_out = time_str
             today_log.timestamp_out = now_utc
             today_log.check_out_status = second["out_status"]
             today_log.check_out_period = second["period"]
             today_log.status = second["final_status"]
             msg = second["message"]
             db.session.commit()
    else:
        # Check in using configured first-mark rules
        first = classify_first_mark(now_local)
        if not first["allowed"]:
            return jsonify({"success": False, "message": first["message"]}), 400

        new_log = AttendanceLog(
            user_id=user_id,
            date=today_str,
            time_in=time_str,
            time_out=None,
            check_in_status=first["status"],
            check_in_period=first["period"],
            timestamp_in=now_utc,
            status=first["status"]
        )
        db.session.add(new_log)
        db.session.commit()
        msg = first["message"]

    # Audit Log
    audit = AuditLog(
        admin_id=admin_id,
        action=f"Marked {log_type} for {user.name} ({user_id})",
        timestamp=now_utc,
        reason="Manual Entry"
    )
    db.session.add(audit)

    return jsonify({
        "success": True,
        "message": msg,
        "entry": {
            "type": log_type,
            "time": time_str,
            "status": "Updated"
        }
    })

# 6. Get Dashboard Data
@app.route('/api/dashboard/<role>/<user_id>', methods=['GET'])
def get_dashboard(role, user_id):
    # Use UTC to filter for "Today" requires conversion
    now_utc = datetime.now(pytz.utc)
    local_tz = pytz.timezone('Asia/Kolkata') # Configurable
    now_local = now_utc.astimezone(local_tz)
    today_str = now_local.strftime('%Y-%m-%d')

    if role == 'admin':
        # Admin sees everything
        # EXCLUDE ADMIN01 from stats
        total = User.query.filter(User.user_id != 'ADMIN01').count()

        # Filter logs to exclude ADMIN01 actions if they exist
        logs_today = AttendanceLog.query.filter(AttendanceLog.date == today_str, AttendanceLog.user_id != 'ADMIN01').all()

        # Unique users present today
        present_user_ids = set([l.user_id for l in logs_today if not is_absence_status(l.status)])
        present_count = len(present_user_ids)

        # Calculate Absent
        absent_count = total - present_count

        # Calculate Late Arrivals
        late_logs = [l for l in logs_today if l.status and ('Late' in l.status)]
        late_count = len(set([l.user_id for l in late_logs]))

        # Recent logs (Exclude ADMIN01)
        recent_logs = AttendanceLog.query.filter(AttendanceLog.user_id != 'ADMIN01').order_by(AttendanceLog.timestamp_in.desc()).limit(10).all()
        logs_data = []
        for log in recent_logs:
            u = User.query.filter_by(user_id=log.user_id).first()
            if u and u.user_id == 'ADMIN01': continue # Double safety check

            logs_data.append({
                "id": log.user_id,
                "name": u.name if u else "Unknown",
                "role": u.role if u else "-",
                "date": log.date,
                "time_in": log.time_in,
                "time_out": log.time_out if log.time_out else "-",
                "status": log.status
            })

        return jsonify({
            "stats": {
                "total_users": total,
                "present_today": present_count,
                "late_count": late_count,
                "absent_count": absent_count
            },
            "logs": logs_data
        })

    elif role == 'faculty' or role == 'student':
        # See only own logs
        # Order by Date descending
        logs = AttendanceLog.query.filter_by(user_id=user_id).order_by(AttendanceLog.id.desc()).all()
        logs_data = []
        for l in logs:
            duration_str = "-"
            if l.timestamp_in and l.timestamp_out:
                diff = l.timestamp_out - l.timestamp_in
                total_seconds = int(diff.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                duration_str = f"{hours}h {minutes}m"

            logs_data.append({
                "date": l.date,
                "time_in": l.time_in,
                "time_out": l.time_out if l.time_out else "-",
                "duration": duration_str,
                "status": l.status
            })

        # Today's Check
        today_log = AttendanceLog.query.filter_by(user_id=user_id, date=today_str).first()

        last_in = "--:--"
        last_out = "--:--"
        current_status = "Not Marked"

        if today_log:
            last_in = today_log.time_in if today_log.time_in else "--:--"
            last_out = today_log.time_out if today_log.time_out else "--:--"

            if today_log.time_out:
                current_status = f"Checked Out ({today_log.status})" if today_log.status else "Checked Out"
            else:
                current_status = today_log.status if today_log.status else "Checked In"

        return jsonify({
            "logs": logs_data,
            "current_status": current_status,
            "last_check_in": last_in,
            "last_check_out": last_out
        })

    return jsonify({"error": "Invalid Role"}), 400

# 7. Bulk Import Users
@app.route('/api/bulk_import', methods=['POST'])
def bulk_import():
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "No file part"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"success": False, "message": "No selected file"}), 400

        if file:
            df = pd.read_csv(file)
            # Expected columns: user_id, name, role, password (optional)

            count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    uid = str(row['user_id'])
                    # Block importing as ADMIN01
                    if uid == 'ADMIN01':
                         continue

                    if User.query.filter_by(user_id=uid).first():
                        errors.append(f"User {uid} already exists")
                        continue

                    raw_pw = row.get('password', uid)
                    pw_hash = bcrypt.hashpw(str(raw_pw).encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

                    new_user = User(
                        user_id=uid,
                        name=row['name'],
                        role=row['role'],
                        password_hash=pw_hash,
                        face_encoding=None # Import text data only, face needs registration later
                    )
                    db.session.add(new_user)
                    count += 1
                except Exception as row_err:
                    errors.append(f"Row {index}: {str(row_err)}")

            db.session.commit()

            return jsonify({
                "success": True,
                "message": f"Imported {count} users.",
                "errors": errors
            })

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# 7. Reports - Analytics Data
@app.route('/api/report', methods=['GET'])
def get_analytics_data():
    try:
        # 1. Get Filters
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        # Default to last 30 days if not provided
        if not start_date_str or not end_date_str:
            today = datetime.now()
            start_date = today - pd.Timedelta(days=30)
            start_date_str = start_date.strftime('%Y-%m-%d')
            end_date_str = today.strftime('%Y-%m-%d')

        # 2. Get Total Users (Excluding Admin) for "Absent" calculation
        relevant_users = User.query.filter(User.user_id != 'ADMIN01').all()
        total_staff_count = len(relevant_users)

        # 3. Fetch Logs in Range
        logs = AttendanceLog.query.filter(
            AttendanceLog.date >= start_date_str,
            AttendanceLog.date <= end_date_str,
            AttendanceLog.user_id != 'ADMIN01'
        ).all()

        # 4. Aggregate Data by Date
        daily_map = {}

        # Initialize map for fill gaps
        try:
            s_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            e_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            delta = e_date - s_date

            for i in range(delta.days + 1):
                day = s_date + pd.Timedelta(days=i)
                day_str = day.strftime('%Y-%m-%d')
                daily_map[day_str] = {
                    "date": day_str,
                    "present": 0,
                    "absent": total_staff_count, # Default all absent
                    "late": 0,
                    "check_ins": []
                }
        except Exception as date_err:
            print(f"Date Parsing Error: {date_err}")

        # Fill with log data
        for log in logs:
            if log.date in daily_map:
                entry = daily_map[log.date]
                if not is_absence_status(log.status):
                    entry["present"] += 1
                    entry["absent"] = max(0, entry["absent"] - 1)

                    if log.status and "Late" in log.status:
                        entry["late"] += 1

                    if log.time_in:
                        try:
                            t = datetime.strptime(log.time_in, "%H:%M:%S")
                            seconds = t.hour * 3600 + t.minute * 60 + t.second
                            entry["check_ins"].append(seconds)
                        except:
                           pass

        # 5. Finalize Daily Stats List
        daily_stats = []
        for date_key in sorted(daily_map.keys()):
            d = daily_map[date_key]

            # Calc Average Time
            avg_time_str = "-"
            if d["check_ins"]:
                avg_sec = sum(d["check_ins"]) / len(d["check_ins"])
                m, s = divmod(avg_sec, 60)
                h, m = divmod(m, 60)
                avg_time_str = "{:02d}:{:02d}".format(int(h), int(m))

            daily_stats.append({
                "date": d["date"],
                "present": d["present"],
                "absent": d["absent"],
                "late": d["late"],
                "avg_check_in": avg_time_str
            })

        # 6. Department Breakdown
        dept_counts = {}
        user_role_map = {u.user_id: u.role for u in relevant_users}
        user_name_map = {u.user_id: u.name for u in relevant_users}

        for log in logs:
            role = user_role_map.get(log.user_id, "Unknown")
            dept_counts[role] = dept_counts.get(role, 0) + 1

        dept_stats = [{"label": k, "value": v} for k, v in dept_counts.items()]

        # 7. Late Arrival Leaderboard (Top 5)
        late_counts = {}
        for log in logs:
            if log.status and "Late" in log.status:
                late_counts[log.user_id] = late_counts.get(log.user_id, 0) + 1

        sorted_late = sorted(late_counts.items(), key=lambda item: item[1], reverse=True)[:5]
        leaderboard = [{"name": user_name_map.get(uid, uid), "count": count} for uid, count in sorted_late]

        # 8. Check-in Time Distribution
        # Define buckets based on system logic (9:30 Start)
        distribution = {
            "Early (Before 9:00)": 0,
            "On Time (9:00 - 9:30)": 0,
            "Early Permission (9:30 - 9:45)": 0,
            "Late (After 9:45)": 0
        }

        for log in logs:
            if not log.time_in: continue
            try:
                t = datetime.strptime(log.time_in, "%H:%M:%S").time()
                # Compare
                if t < time(9, 0):
                    distribution["Early (Before 9:00)"] += 1
                elif t < time(9, 30):
                    distribution["On Time (9:00 - 9:30)"] += 1
                elif t <= time(9, 45):
                    distribution["Early Permission (9:30 - 9:45)"] += 1
                else:
                    distribution["Late (After 9:45)"] += 1
            except:
                pass

        # 9. Recent Check-ins (Limit 5)
        # We need to re-query for sorted logs or sort the current list
        # detailed logs are already in 'logs', but not sorted by time necessarily (db default?)
        # Let's sort manually
        sorted_logs = sorted(logs, key=lambda x: (x.date, x.time_in if x.time_in else ""), reverse=True)[:5]
        recent_activity = []
        for log in sorted_logs:
            recent_activity.append({
                "name": user_name_map.get(log.user_id, log.user_id),
                "time": f"{log.date} {log.time_in}" if log.time_in else f"{log.date} (No Time)",
                "status": log.status
            })

        return jsonify({
            "daily_stats": daily_stats,
            "department_stats": dept_stats,
            "leaderboard": leaderboard,
            "time_distribution": distribution,
            "recent_activity": recent_activity,
            "summary": {
                "total_days": len(daily_stats),
                "total_present_events": sum(d['present'] for d in daily_stats),
                "total_faculty": total_staff_count
            }
        })

    except Exception as e:
        print(f"Analytics Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/download_report', methods=['GET'])
def download_excel():
    user_id = request.args.get('user_id')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    report_type = request.args.get('type', 'detailed') # 'detailed' or 'summary'

    if not start_date_str or not end_date_str:
        return jsonify({"error": "Start and End dates are required."}), 400

    try:
        # Fetch logs
        query = AttendanceLog.query.filter(
            AttendanceLog.date >= start_date_str,
            AttendanceLog.date <= end_date_str,
            AttendanceLog.user_id != 'ADMIN01' # EXCLUDE ADMIN
        )
        if user_id:
             query = query.filter_by(user_id=user_id)

        # Sort by Date descending (newest first)
        logs = query.order_by(AttendanceLog.date.desc(), AttendanceLog.time_in.desc()).all()

        if report_type == 'detailed':
            # --- 1. DETAILED MONTHLY REPORT ---
            # Columns: Staff ID, Name, Position, Date, Check In, Check Out, Check In Status, Check Out Status

            report_data = []
            for log in logs:
                u = User.query.filter_by(user_id=log.user_id).first()
                status_in = log.check_in_status if hasattr(log, 'check_in_status') and log.check_in_status else (log.status if log.time_in else "-")
                status_out = log.check_out_status if hasattr(log, 'check_out_status') and log.check_out_status else (log.status if log.time_out else "-")

                report_data.append({
                    "Staff ID": log.user_id,
                    "Name": u.name if u else "Unknown",
                    "Position": u.role if u else "-",
                    "Date": log.date,
                    "Check In": log.time_in if log.time_in else "-",
                    "Check Out": log.time_out if log.time_out else "-",
                    "Check In Status": status_in,
                    "Check Out Status": status_out
                })

            df = pd.DataFrame(report_data)
            filename = f"attendance_report_{start_date_str[:7]}.csv" # YYYY-MM

        elif report_type == 'summary':
            # --- 2. MONTHLY SUMMARY REPORT ---
            # Aggr per user.
            # Columns: Staff ID, Name, Position, Days Present, Working Days, Attendance %, Status

            # Get all users (filtered if user_id provided)
            if user_id:
                users_list = [User.query.filter_by(user_id=user_id).first()]
            else:
                users_list = User.query.filter(User.user_id != 'ADMIN01').all() # EXCLUDE ADMIN

            report_data = []

            # Calculate Working Days in Range
            # "Total weekdays - Holidays"
            # 1. Parse dates
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")

            # 2. Count weekdays
            total_days = (end_dt - start_dt).days + 1
            weekdays_count = 0
            for i in range(total_days):
                day = start_dt + pd.Timedelta(days=i)
                if day.weekday() < 5: # 0-4 is Mon-Fri
                    weekdays_count += 1

            # 3. Subtract Holidays
            holidays_count = Holiday.query.filter(Holiday.date >= start_date_str, Holiday.date <= end_date_str).count()
            working_days = max(1, weekdays_count - holidays_count) # Avoid division by zero

            for u in users_list:
                if not u: continue
                # Count days present for this user in this range
                days_present = AttendanceLog.query.filter(
                    AttendanceLog.user_id == u.user_id,
                    AttendanceLog.date >= start_date_str,
                    AttendanceLog.date <= end_date_str,
                    AttendanceLog.time_in != None # Must have at least check in
                ).count()

                raw_pct = (days_present / working_days) * 100
                attendance_pct = round(raw_pct, 1)

                # Categorize
                perf_status = "Poor"
                if attendance_pct >= 80:
                    perf_status = "Excellent"
                elif attendance_pct >= 50:
                    perf_status = "Good"

                report_data.append({
                    "Staff ID": u.user_id,
                    "Name": u.name,
                    "Position": u.role,
                    "Days Present": days_present,
                    "Working Days": working_days,
                    "Attendance %": f"{attendance_pct}%",
                    "Status": perf_status
                })

            df = pd.DataFrame(report_data)
            filename = f"attendance_summary_{start_date_str[:7]}.csv" # YYYY-MM

        else:
            return jsonify({"error": "Invalid report type"}), 400

        # Export CSV
        output = io.BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)

        response = make_response(send_file(
            output,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename
        ))
        # Mobile-compatibility: expose Content-Disposition so JS can read filename
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response.headers['Cache-Control'] = 'no-store'
        return response

    except Exception as e:
        print("Report Gen Error:", e)
        return jsonify({"error": str(e)}), 500


def filter_report_columns(report_data, report_type, selected_columns):
    """
    Filter report data to include only selected columns based on user preferences.

    Args:
        report_data: List of dictionaries containing report
        report_type: Type of report (summary, detailed, violations, compliance)
        selected_columns: Dict with keys {user, date, status, location, checkin, period}

    Returns:
        Filtered list of dictionaries with only selected columns
    """
    if not report_data or not selected_columns:
        return report_data

    # Define column mappings for each report type
    column_map = {
        'summary': {
            'user': ['Staff ID', 'Name', 'Role'],
            'date': [],
            'status': ['Present Days', 'Absent Days', 'Late Days', 'Attendance %'],
            'location': [],
            'checkin': [],
            'period': ['Attendance %']
        },
        'detailed': {
            'user': ['Staff ID', 'Name', 'Role'],
            'date': ['Date'],
            'status': ['Day Status', 'Check-in Status', 'Check-out Status'],
            'location': [],
            'checkin': ['Check-in', 'Check-out'],
            'period': ['Period']
        },
        'violations': {
            'user': ['Staff ID', 'Name'],
            'date': ['Date'],
            'status': ['Violation Type'],
            'location': [],
            'checkin': ['Check-in'],
            'period': ['Severity']
        },
        'compliance': {
            'user': ['Staff ID', 'Name', 'Role'],
            'date': [],
            'status': ['Compliance %', 'Status'],
            'location': [],
            'checkin': [],
            'period': ['Working Days', 'Present', 'Late', 'Absent']
        }
    }

    # Device state report has its own column structure
    if report_type == 'device_state':
        return report_data  # No column filtering needed — all columns are always relevant

    # Always include core identifying columns
    always_include = set()
    if report_type != 'violations':  # Violations only has Staff ID and Name
        always_include = {'Staff ID', 'Name', 'Role'} if report_type != 'violations' else {'Staff ID', 'Name'}

    # Build list of columns to keep
    cols_to_keep = always_include.copy()

    if report_type in column_map:
        for col_key, col_list in column_map[report_type].items():
            if selected_columns.get(col_key, False):
                cols_to_keep.update(col_list)

    # Filter each row to only include selected columns
    if report_data and isinstance(report_data[0], dict):
        filtered_data = []
        for row in report_data:
            # Only include columns that both exist in the row AND are in cols_to_keep
            filtered_row = {k: v for k, v in row.items() if k in cols_to_keep}
            filtered_data.append(filtered_row)
        return filtered_data

    return report_data


# Export Report API (Multiple Formats)
@app.route('/api/export_report', methods=['GET'])
def export_report():
    """
    Advanced export endpoint supporting CSV, Excel, PDF formats and various filters.
    """
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        report_type = request.args.get('type', 'summary')
        file_format = request.args.get('format', 'csv')
        filter_latest = request.args.get('filter_latest', 'true').lower() == 'true'
        filter_violations = request.args.get('filter_violations', 'false').lower() == 'true'
        include_timestamps = request.args.get('include_timestamps', 'true').lower() == 'true'

        # Parse column preferences from frontend
        columns_param = request.args.get('columns', '{}')
        try:
            selected_columns = json.loads(columns_param)
        except:
            selected_columns = {}

        if not start_date_str or not end_date_str:
            return jsonify({"error": "Start and End dates required"}), 400

        # Validate format
        if file_format not in ['csv', 'excel', 'pdf', 'json']:
            file_format = 'csv'

        # Parse dates safely
        try:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400

        # Build base query
        base_query = AttendanceLog.query.filter(
            AttendanceLog.date >= start_date_str,
            AttendanceLog.date <= end_date_str,
            AttendanceLog.user_id != 'ADMIN01'
        )

        # Apply violations filter
        if filter_violations:
            base_query = base_query.filter(
                (AttendanceLog.time_in == None) |
                (AttendanceLog.status.in_(['Late', 'Absent', 'Out-of-Bounds']))
            )

        # Fetch logs
        logs = base_query.order_by(AttendanceLog.date.desc(), AttendanceLog.timestamp_in.desc()).all()

        # Apply latest filter
        if filter_latest and logs:
            seen = {}
            unique_logs = []
            for log in logs:
                key = (log.user_id, log.date)
                if key not in seen:
                    seen[key] = True
                    unique_logs.append(log)
            logs = unique_logs

        # Generate report based on type
        if report_type == 'summary':
            report_data = generate_summary_report(logs, start_dt, end_dt, include_timestamps)
        elif report_type == 'detailed':
            report_data = generate_detailed_report(logs, include_timestamps)
        elif report_type == 'violations':
            report_data = generate_violations_report(logs, include_timestamps)
        elif report_type == 'compliance':
            report_data = generate_compliance_report(logs, start_dt, end_dt)
        elif report_type == 'device_state':
            # Device state report uses its own query (DeviceStateLog, not AttendanceLog)
            report_data = generate_device_state_report(start_date_str, end_date_str)
        else:
            return jsonify({"error": "Invalid report type"}), 400

        if not report_data:
            report_data = []

        # Apply column filtering based on user selection
        report_data = filter_report_columns(report_data, report_type, selected_columns)

        # JSON format - return as JSON preview
        if file_format == 'json':
            return jsonify({
                "status": "success",
                "report_type": report_type,
                "date_range": f"{start_date_str} to {end_date_str}",
                "record_count": len(report_data),
                "data": report_data[:10]  # Limit to 10 for preview
            })

        # CSV format
        if file_format == 'csv':
            try:
                df = pd.DataFrame(report_data) if report_data else pd.DataFrame()
                output = io.BytesIO()
                df.to_csv(output, index=False)
                output.seek(0)

                return send_file(
                    output,
                    mimetype="text/csv",
                    as_attachment=True,
                    download_name=f"FaceAttend_Report_{report_type}_{start_date_str}.csv"
                )
            except Exception as e:
                return jsonify({"error": f"CSV export failed: {str(e)}"}), 500

        # Excel format
        if file_format == 'excel':
            try:
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                df = pd.DataFrame(report_data) if report_data else pd.DataFrame()
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Report')

                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Report']

                    # Header formatting
                    header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF', size=11)
                    center_align = Alignment(horizontal='center', vertical='center')
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Format header row
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                        cell.border = border

                    # Alternate row colors and borders
                    light_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
                    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                        for cell in row:
                            cell.border = border
                            if row_num % 2 == 0:
                                cell.fill = light_fill
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                output.seek(0)

                return send_file(
                    output,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name=f"FaceAttend_Report_{report_type}_{start_date_str}.xlsx"
                )
            except Exception as e:
                return jsonify({"error": f"Excel export failed: {str(e)}"}), 500

        # PDF format
        if file_format == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, PageBreak
                from reportlab.lib.units import inch
                from reportlab.lib.enums import TA_CENTER, TA_LEFT

                df = pd.DataFrame(report_data) if report_data else pd.DataFrame()
                output = io.BytesIO()
                doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
                elements = []

                # Styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading2'],
                    fontSize=16,
                    textColor=colors.HexColor('#0f172a'),
                    spaceAfter=6,
                    alignment=TA_LEFT,
                    fontName='Helvetica-Bold'
                )
                subtitle_style = ParagraphStyle(
                    'Subtitle',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#64748b'),
                    spaceAfter=12,
                    alignment=TA_LEFT
                )

                # Title and metadata
                title = Paragraph(f"FaceAttend - {report_type.title()} Report", title_style)
                elements.append(title)
                subtitle = Paragraph(f"Date Range: {start_date_str} to {end_date_str} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style)
                elements.append(subtitle)
                elements.append(Spacer(1, 0.2 * inch))

                # Table
                if not df.empty:
                    # Convert all values to strings to avoid reportlab issues with numpy types
                    data = [df.columns.tolist()] + df.astype(str).values.tolist()
                    col_count = len(df.columns)
                    col_width = 7 * inch / col_count if col_count > 0 else 1 * inch
                    t = Table(data, colWidths=[col_width] * col_count)

                    # Enhanced table styling with alternating rows
                    style_commands = [
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                        ('TOPPADDING', (0, 0), (-1, 0), 10),
                        ('PADDING', (0, 1), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
                    ]
                    t.setStyle(TableStyle(style_commands))
                    elements.append(t)

                    # Add summary line
                    elements.append(Spacer(1, 0.2 * inch))
                    summary = Paragraph(f"<i>Total Records: {len(df)} | Export generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</i>", styles['Normal'])
                    elements.append(summary)
                else:
                    elements.append(Paragraph("No data found for the selected date range and filters.", styles['Normal']))

                doc.build(elements)
                output.seek(0)

                return send_file(
                    output,
                    mimetype="application/pdf",
                    as_attachment=True,
                    download_name=f"FaceAttend_Report_{report_type}_{start_date_str}.pdf"
                )
            except Exception as e:
                return jsonify({"error": f"PDF export failed: {str(e)}"}), 500

        return jsonify({"error": "Invalid format"}), 400

    except Exception as e:
        print(f"Export Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


def generate_summary_report(logs, start_dt, end_dt, include_timestamps=True):
    """Generate summary-level attendance report"""
    users_data = {}

    for log in logs:
        user = User.query.filter_by(user_id=log.user_id).first()
        if not user:
            continue

        if log.user_id not in users_data:
            users_data[log.user_id] = {
                'user_id': log.user_id,
                'name': user.name,
                'role': user.role,
                'days_present': 0,
                'days_absent': 0,
                'days_late': 0,
                'total_records': 0
            }

        users_data[log.user_id]['total_records'] += 1

        # Safely check status
        status = log.status if log.status else 'Unknown'
        has_time_in = log.time_in is not None and str(log.time_in).strip() != ''

        if status == 'Absent' or not has_time_in:
            users_data[log.user_id]['days_absent'] += 1
        elif status == 'Late Permission':
            users_data[log.user_id]['days_late'] += 1
        else:
            users_data[log.user_id]['days_present'] += 1

    # Calculate working days
    total_days = (end_dt - start_dt).days + 1
    weekdays = [i for i in range(total_days) if (start_dt + pd.Timedelta(days=i)).weekday() < 5]
    weekdays_count = len(weekdays)

    holidays = Holiday.query.filter(
        Holiday.date >= start_dt.strftime('%Y-%m-%d'),
        Holiday.date <= end_dt.strftime('%Y-%m-%d')
    ).count()
    working_days = max(1, weekdays_count - holidays)

    # Build report
    report_data = []
    for uid, data in users_data.items():
        present = data['days_present']
        attendance_pct = round((present / working_days * 100), 1) if working_days > 0 else 0

        row = {
            'Staff ID': uid,
            'Name': data['name'],
            'Role': data['role'],
            'Present Days': present,
            'Absent Days': data['days_absent'],
            'Late Days': data['days_late'],
            'Attendance %': f"{attendance_pct}%"
        }
        report_data.append(row)

    return report_data


def generate_detailed_report(logs, include_timestamps=True):
    """Generate detailed attendance log report"""
    report_data = []
    for log in logs:
        user = User.query.filter_by(user_id=log.user_id).first()

        # Determine check-in status based on actual check-in
        if log.time_in:
            # User checked in - use check_in_status or fall back to check calendar-based status (P/LP/HD/EP)
            status_in = log.check_in_status if (hasattr(log, 'check_in_status') and log.check_in_status) else 'Present'
        else:
            # No check-in time recorded
            status_in = 'Absent'

        status_out = log.check_out_status if hasattr(log, 'check_out_status') and log.check_out_status else '-'

        row = {
            'Date': log.date,
            'Staff ID': log.user_id,
            'Name': user.name if user else 'Unknown',
            'Role': user.role if user else 'N/A',
            'Check-in': log.time_in if log.time_in else 'Absent',
            'Check-out': log.time_out or '-',
            'Check-in Status': status_in,
            'Check-out Status': status_out,
            'Day Status': log.status,
            'Period': f"{log.check_in_period or '-'} / {log.check_out_period or '-'}"
        }
        report_data.append(row)

    return report_data


def generate_violations_report(logs, include_timestamps=True):
    """Generate violations-only report (absences, lates, out-of-bounds)"""
    report_data = []
    for log in logs:
        # Skip non-violation day statuses
        if log.status in ['FD', 'Present', 'Late Permission', 'HD']:
            continue

        user = User.query.filter_by(user_id=log.user_id).first()
        violation_type = 'Absent' if not log.time_in else (log.status or 'Other')

        row = {
            'Date': log.date,
            'Staff ID': log.user_id,
            'Name': user.name if user else 'Unknown',
            'Violation Type': violation_type,
            'Check-in': log.time_in or '-',
            'Severity': 'High' if violation_type == 'Absent' else 'Medium'
        }
        if include_timestamps:
            row['Timestamp'] = log.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(log, 'created_at') else '-'
        report_data.append(row)

    return report_data


def generate_compliance_report(logs, start_dt, end_dt):
    """Generate compliance metrics report"""
    # Calculate working days
    total_days = (end_dt - start_dt).days + 1
    weekdays_count = sum(1 for i in range(total_days)
                        if (start_dt + pd.Timedelta(days=i)).weekday() < 5)
    holidays = Holiday.query.filter(
        Holiday.date >= start_dt.strftime('%Y-%m-%d'),
        Holiday.date <= end_dt.strftime('%Y-%m-%d')
    ).count()
    working_days = max(1, weekdays_count - holidays)

    # Get all unique users
    all_users = User.query.filter(User.user_id != 'ADMIN01').all()

    report_data = []
    for user in all_users:
        # Count records for this user in date range
        user_logs = [l for l in logs if l.user_id == user.user_id]

        present_days = sum(1 for l in user_logs if l.status in ['FD', 'HD', 'Present', 'Late Permission', 'EP'])
        late_days = sum(1 for l in user_logs if (hasattr(l, 'check_in_status') and l.check_in_status == 'Late Permission'))
        absent_days = sum(1 for l in user_logs if l.status == 'Absent' or not l.time_in)

        compliance_pct = round((present_days / working_days * 100), 1) if working_days > 0 else 0
        compliance_status = 'Compliant' if compliance_pct >= 80 else ('Non-Compliant' if compliance_pct < 60 else 'At Risk')

        row = {
            'Staff ID': user.user_id,
            'Name': user.name,
            'Role': user.role,
            'Working Days': working_days,
            'Present': present_days,
            'Late': late_days,
            'Absent': absent_days,
            'Compliance %': f"{compliance_pct}%",
            'Status': compliance_status
        }
        report_data.append(row)

    return report_data


def generate_device_state_report(start_date_str, end_date_str, user_id_filter=None):
    """Generate a comprehensive state-transition timeline from DeviceStateLog.
    Each row is one state transition event with duration calculations.
    Shows ALL dimensions: Network, GPS Location, Campus Bounds.
    """
    query = DeviceStateLog.query.filter(
        DeviceStateLog.date >= start_date_str,
        DeviceStateLog.date <= end_date_str
    )
    if user_id_filter:
        query = query.filter_by(user_id=user_id_filter)

    logs = query.order_by(DeviceStateLog.user_id, DeviceStateLog.date.asc(), DeviceStateLog.timestamp.asc()).all()

    # Build user name lookup
    user_ids = list(set(log.user_id for log in logs))
    users = {u.user_id: u for u in User.query.filter(User.user_id.in_(user_ids)).all()} if user_ids else {}

    # Group by user for duration calculation
    from collections import defaultdict
    user_logs = defaultdict(list)
    for log in logs:
        user_logs[log.user_id].append(log)

    report_data = []
    for uid, entries in user_logs.items():
        user = users.get(uid)
        for i, log in enumerate(entries):
            # Convert UTC timestamp to IST
            ist_time = ''
            ist_datetime = ''
            if log.timestamp:
                ist_dt = log.timestamp.replace(tzinfo=pytz.utc).astimezone(pytz.timezone('Asia/Kolkata'))
                ist_time = ist_dt.strftime('%I:%M:%S %p')
                ist_datetime = ist_dt.strftime('%Y-%m-%d %I:%M:%S %p')

            # Calculate how long this state lasted (until next event of same type)
            duration_str = '-'
            if log.timestamp:
                # Find next event of same type for same user
                next_event = None
                for j in range(i + 1, len(entries)):
                    if entries[j].event_type == log.event_type:
                        next_event = entries[j]
                        break
                if next_event and next_event.timestamp:
                    delta = next_event.timestamp - log.timestamp
                    total_secs = int(delta.total_seconds())
                    if total_secs < 0:
                        duration_str = '-'
                    elif total_secs < 60:
                        duration_str = f'{total_secs}s'
                    elif total_secs < 3600:
                        duration_str = f'{total_secs // 60}m {total_secs % 60}s'
                    else:
                        hrs = total_secs // 3600
                        mins = (total_secs % 3600) // 60
                        duration_str = f'{hrs}h {mins}m'
                else:
                    duration_str = 'Ongoing'

            # Human-readable labels with status indicators
            event_icons = {'NETWORK': '📡', 'LOCATION': '📍', 'BOUNDS': '🏫'}
            state_labels = {
                'NETWORK': {'online': '🟢 Online', 'offline': '🔴 Offline'},
                'LOCATION': {'active': '🟢 GPS Active', 'inactive': '🔴 GPS Off'},
                'BOUNDS': {'in': '🟢 Inside Campus', 'out': '🔴 Outside Campus'}
            }

            icon = event_icons.get(log.event_type, '❓')
            event_label = {'NETWORK': 'Network', 'LOCATION': 'GPS Location', 'BOUNDS': 'Campus Boundary'}.get(log.event_type, log.event_type)
            
            old_labels = state_labels.get(log.event_type, {})
            new_labels = state_labels.get(log.event_type, {})
            old_label = old_labels.get(log.old_state, log.old_state)
            new_label = new_labels.get(log.new_state, log.new_state)

            # Determine if this is a problem state
            problem_states = {'offline', 'inactive', 'out'}
            is_problem = log.new_state in problem_states
            severity = 'Critical' if log.new_state in ('offline',) else ('Warning' if is_problem else 'Normal')

            row = {
                'Date': log.date,
                'Time (IST)': ist_time,
                'Full Timestamp': ist_datetime,
                'Staff ID': log.user_id,
                'Name': user.name if user else 'Unknown',
                'Category': f'{icon} {event_label}',
                'Event': f'{event_label}: {log.old_state} → {log.new_state}',
                'Previous State': old_label,
                'New State': new_label,
                'Duration in State': duration_str,
                'Severity': severity,
                'Distance from Campus (m)': round(log.distance_m, 1) if log.distance_m is not None else '-',
                'GPS Accuracy (m)': round(log.accuracy_m, 1) if log.accuracy_m is not None else '-',
                'Latitude': round(log.latitude, 6) if log.latitude is not None else '-',
                'Longitude': round(log.longitude, 6) if log.longitude is not None else '-',
            }
            report_data.append(row)

    # Sort final output by date desc, time desc for display
    report_data.sort(key=lambda r: r.get('Full Timestamp', ''), reverse=True)

    return report_data

# 9. Get All Users (For Staff Directory)
@app.route('/api/users', methods=['GET'])
def get_users():
    # EXCLUDE ADMIN
    users = User.query.filter(User.user_id != 'ADMIN01').all()
    user_list = []
    today = datetime.now().strftime('%Y-%m-%d')

    for u in users:
        # Check today's status
        log = AttendanceLog.query.filter_by(user_id=u.user_id, date=today).first()
        status = log.status if log else "Absent" # Default to Absent if no log

        user_list.append({
            "id": u.user_id,
            "name": u.name,
            "role": u.role,
            "status": status
        })
    return jsonify(user_list)

# 9. Get All Faculty Users
@app.route('/api/users/faculty', methods=['GET'])
def get_faculty_users():
    """Get all faculty users (for admin messaging)"""
    try:
        faculty = User.query.filter(
            db.func.lower(User.role) == 'faculty',
            User.is_active == True
        ).order_by(User.name.asc()).all()
        faculty_list = [{
            "user_id": f.user_id,
            "name": f.name,
            "email": f.email,
            "role": f.role
        } for f in faculty]

        return jsonify({"success": True, "faculty": faculty_list})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# 9A. Get Single User Details
@app.route('/api/users/<user_id>', methods=['GET'])
def get_user_details(user_id):
    """Get details of a specific user"""
    try:
        user = User.query.filter_by(user_id=user_id).first()
        if not user:
            return jsonify({"success": False, "message": "User not found"}), 404

        today = datetime.now().strftime('%Y-%m-%d')
        log = AttendanceLog.query.filter_by(user_id=user_id, date=today).first()
        status = log.status if log else "Absent"

        return jsonify({
            "success": True,
            "user": {
                "id": user.user_id,
                "name": user.name,
                "role": user.role,
                "status": status,
                "registration_date": user.registration_date.strftime('%Y-%m-%d') if user.registration_date else None
            }
        }), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# 9B. Delete User
@app.route('/api/users/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    try:
        user = User.query.filter_by(user_id=user_id).first()
        if not user:
            return jsonify({"success": False, "message": "User not found"}), 404

        # Delete related logs first
        AttendanceLog.query.filter_by(user_id=user_id).delete()
        PermissionRequest.query.filter_by(user_id=user_id).delete()
        LivePresence.query.filter_by(user_id=user_id).delete() # <-- ADD THIS LINE

        # Delete user
        db.session.delete(user)
        db.session.commit()

        return jsonify({"success": True, "message": f"User {user_id} deleted successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

# 10. Mark User as Absent / Didn't Mark
@app.route('/api/mark_absent', methods=['POST'])
def mark_absent():
    try:
        data = request.json or {}
        user_id = data.get('user_id')
        date = data.get('date') or datetime.now().strftime('%Y-%m-%d')
        admin_id = (data.get('admin_id') or '').strip()

        admin_user = require_active_admin(admin_id)
        if not admin_user:
            return jsonify({"success": False, "message": "Unauthorized. Active admin required."}), 403

        # Verify user exists
        user = User.query.filter_by(user_id=user_id).first()
        if not user:
            return jsonify({"success": False, "message": "User not found"}), 404

        # Check if attendance log exists for the date
        today_log = AttendanceLog.query.filter_by(user_id=user_id, date=date).first()

        if today_log:
            # Update existing log
            today_log.status = "Didn't Mark"
            db.session.commit()
        else:
            # Create new log marked as didn't mark
            new_log = AttendanceLog(
                user_id=user_id,
                date=date,
                status="Didn't Mark"
            )
            db.session.add(new_log)
            db.session.commit()

        # Log admin action
        audit = AdminAuditLog(
            action='marked_absent',
            admin_id=admin_id,
            target_id=user_id,
            description=f"Marked {user.name} as 'Didn't Mark' for {date}"
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({"success": True, "message": f"{user.name} marked as 'Didn't Mark'"}), 200
    except Exception as e:
        print(f"Mark Absent Error: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

if __name__ == '__main__':
    initialize_database_on_startup()

# --------- COMPREHENSIVE NOTIFICATION API ---------

@app.route('/api/user/notifications', methods=['GET'])
def get_user_notifications():
    """
    Get comprehensive notifications for a user
    Returns: Active alerts, warnings, and policy status
    """
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({"success": False, "message": "user_id required"}), 400

    user = User.query.filter_by(user_id=user_id).first()
    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404

    notifications = []
    now_utc = datetime.utcnow()
    local_tz = pytz.timezone('Asia/Kolkata')
    now_local = now_utc.astimezone(local_tz)
    today_str = now_local.strftime('%Y-%m-%d')
    flags = get_time_policy_flags(now_local)

    # 1. Check Network Status (simulate from heartbeat)
    presence = LivePresence.query.filter_by(user_id=user_id).first()

    if presence and presence.status_code == 'NETWORK_OFF':
        notifications.append({
            'type': 'NETWORK_OFF',
            'level': 'CRITICAL',
            'title': '📡 Network Offline',
            'message': 'Mobile data or Wi-Fi is currently OFF. Turn on to continue tracking.',
            'timestamp': now_local.isoformat(),
            'dismissible': False
        })

    # 2. Check Location Status
    if presence and presence.status_code == 'LOCATION_OFF':
        notifications.append({
            'type': 'LOCATION_OFF',
            'level': 'CRITICAL',
            'title': '📍 Location Services Disabled',
            'message': 'GPS/Location permission is OFF. Enable GPS to continue attendance tracking.',
            'timestamp': now_local.isoformat(),
            'dismissible': False
        })

    # 3. Check Out of Bounds Status
    if presence and presence.status_code == 'OUT_OF_BOUNDS' and not flags['is_lunch_window']:
        distance_str = f"{presence.distance_m:.0f}m" if presence.distance_m else "unknown"
        notifications.append({
            'type': 'OUT_OF_BOUNDS',
            'level': 'WARNING',
            'title': '⚠️ Outside Campus Bounds',
            'message': f'You are {distance_str} outside campus. Return immediately to avoid absent marking.',
            'timestamp': now_local.isoformat(),
            'dismissible': True,
            'distance_m': presence.distance_m
        })

    # 4. Lunch Period Alerts
    if flags['is_lunch_pre_alert']:
        notifications.append({
            'type': 'LUNCH_START_REMINDER',
            'level': 'INFO',
            'title': '🍽️ Lunch Break Starting Soon',
            'message': 'Lunch window starts at 1:00 PM. Lunch free-exit: 1:00 PM - 1:40 PM.',
            'timestamp': now_local.isoformat(),
            'dismissible': True
        })

    if flags['is_lunch_window']:
        notifications.append({
            'type': 'LUNCH_ACTIVE',
            'level': 'INFO',
            'title': '🍽️ Lunch Break Active',
            'message': 'You can exit campus during lunch. Return by 1:40 PM.',
            'timestamp': now_local.isoformat(),
            'dismissible': False
        })

    if flags['is_return_pre_alert']:
        notifications.append({
            'type': 'LUNCH_END_REMINDER',
            'level': 'WARNING',
            'title': '⏰ Lunch Ending Soon',
            'message': '10-minute warning: Lunch free-exit ends at 1:40 PM. Return to campus now!',
            'timestamp': now_local.isoformat(),
            'dismissible': True
        })

    # 5. Check Attendance Status Today
    today_log = AttendanceLog.query.filter_by(user_id=user_id, date=today_str).first()
    if today_log:
        if today_log.time_in and not today_log.time_out:
            notifications.append({
                'type': 'CHECKED_IN',
                'level': 'INFO',
                'title': '✅ Checked In',
                'message': f'You checked in at {today_log.time_in} as {today_log.check_in_status}',
                'timestamp': now_local.isoformat(),
                'dismissible': True
            })
        elif today_log.time_in and today_log.time_out:
            notifications.append({
                'type': 'FULL_DAY_MARKED',
                'level': 'SUCCESS',
                'title': '✨ Full Day Marked',
                'message': f'Check-in: {today_log.time_in} | Check-out: {today_log.time_out} | Status: {today_log.status}',
                'timestamp': now_local.isoformat(),
                'dismissible': True
            })
    else:
        # Not marked
        if flags['is_check_in_window']:
            notifications.append({
                'type': 'PENDING_CHECKIN',
                'level': 'INFO',
                'title': '📸 Mark Attendance',
                'message': 'Check-in window is open (9:00 AM - 9:45 AM). Mark your attendance now.',
                'timestamp': now_local.isoformat(),
                'dismissible': False
            })

    # 6. Check Permission Status
    permissions = PermissionRequest.query.filter_by(
        user_id=user_id,
        date=today_str,
        status='Approved'
    ).all()

    if permissions:
        perm_types = [p.type for p in permissions]
        notifications.append({
            'type': 'APPROVED_PERMISSIONS',
            'level': 'SUCCESS',
            'title': '✅ Permissions Approved',
            'message': f'You have approved permissions: {", ".join(perm_types)}',
            'timestamp': now_local.isoformat(),
            'dismissible': True
        })

    # 7. Check Security Alerts
    recent_alerts = SecurityAlert.query.filter_by(user_id=user_id).filter(
        SecurityAlert.created_at >= (now_utc - timedelta(hours=2))
    ).all()

    if recent_alerts:
        for alert in recent_alerts[:3]:  # Limit to 3 most recent
            notifications.append({
                'type': 'SECURITY_ALERT',
                'level': 'WARNING',
                'title': '🚨 Security Alert',
                'message': alert.event_message,
                'timestamp': alert.created_at.astimezone(local_tz).isoformat(),
                'dismissible': True,
                'code': alert.event_code
            })

    return jsonify({
        "success": True,
        "user_id": user_id,
        "timestamp": now_local.isoformat(),
        "timezone": "Asia/Kolkata",
        "notifications": notifications,
        "stats": {
            "critical": len([n for n in notifications if n['level'] == 'CRITICAL']),
            "warning": len([n for n in notifications if n['level'] == 'WARNING']),
            "info": len([n for n in notifications if n['level'] == 'INFO']),
            "success": len([n for n in notifications if n['level'] == 'SUCCESS']),
            "total": len(notifications)
        }
    }), 200

# --------- FCM TOKEN MANAGEMENT ---------

@app.route('/api/fcm/register_token', methods=['POST'])
def register_fcm_token():
    """Register or update FCM token for a user"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        fcm_token = data.get('fcm_token')
        device_info = data.get('device_info', 'Unknown Device')

        if not user_id or not fcm_token:
            return jsonify({"error": "Missing user_id or fcm_token"}), 400

        user = User.query.filter_by(user_id=user_id).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Check if token already exists
        existing_token = FCMToken.query.filter_by(fcm_token=fcm_token).first()
        if existing_token:
            existing_token.is_active = True
            existing_token.updated_at = datetime.utcnow()
        else:
            new_token = FCMToken(
                user_id=user_id,
                fcm_token=fcm_token,
                device_info=device_info,
                is_active=True
            )
            db.session.add(new_token)

        db.session.commit()
        return jsonify({"success": True, "message": "FCM token registered"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/fcm/tokens/<user_id>', methods=['GET'])
def get_user_fcm_tokens(user_id):
    """Get all FCM tokens for a user"""
    try:
        tokens = FCMToken.query.filter_by(user_id=user_id, is_active=True).all()
        return jsonify([{
            "id": t.id,
            "fcm_token": t.fcm_token[:10] + "...",
            "device_info": t.device_info,
            "created_at": t.created_at.isoformat()
        } for t in tokens]), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --------- IN-APP MESSAGING ---------

@app.route('/api/messages/send', methods=['POST'])
def send_message():
    """Send a message - admins can broadcast, anyone can send direct messages"""
    try:
        data = request.get_json()
        sender_id = data.get('sender_id')
        recipient_id = data.get('recipient_id')
        title = data.get('title')
        content = data.get('content')
        is_broadcast = data.get('is_broadcast', False)

        if not sender_id or not title or not content:
            return jsonify({"error": "Missing required fields"}), 400

        sender = User.query.filter_by(user_id=sender_id).first()
        if not sender:
            return jsonify({"error": "Sender not found"}), 404

        # Only admins can broadcast
        if is_broadcast and sender.role != 'admin':
            return jsonify({"error": "Only admins can broadcast messages"}), 403

        # Create message
        message = Message(
            sender_id=sender_id,
            recipient_id=recipient_id if not is_broadcast else None,
            title=title,
            content=content,
            type='announcement' if is_broadcast else 'message',
            is_broadcast=is_broadcast
        )
        db.session.add(message)
        db.session.flush()

        # If broadcast, create MessageRead entries for all faculty
        if is_broadcast:
            faculty_users = User.query.filter_by(role='faculty', is_active=True).all()
            for faculty in faculty_users:
                msg_read = MessageRead(
                    message_id=message.id,
                    user_id=faculty.user_id,
                    is_read=False
                )
                db.session.add(msg_read)

                # Send FCM wakeup ping to all devices (Admin SDK)
                tokens = FCMToken.query.filter_by(user_id=faculty.user_id, is_active=True).all()
                for token in tokens:
                    ping_device_via_fcm(token.fcm_token, "wakeup")
        else:
            # Direct message
            if not recipient_id:
                return jsonify({"error": "Recipient required for direct messages"}), 400

            recipient = User.query.filter_by(user_id=recipient_id).first()
            if not recipient:
                return jsonify({"error": "Recipient not found"}), 404

            msg_read = MessageRead(
                message_id=message.id,
                user_id=recipient_id,
                is_read=False
            )
            db.session.add(msg_read)

            # Send FCM wakeup ping to recipient (Admin SDK)
            tokens = FCMToken.query.filter_by(user_id=recipient_id, is_active=True).all()
            for token in tokens:
                ping_device_via_fcm(token.fcm_token, "wakeup")

        db.session.commit()
        return jsonify({"success": True, "message_id": message.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/messages/<user_id>', methods=['GET'])
def get_user_messages(user_id):
    """Get all messages for a user (as recipient or from direct messages)"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = 20

        # Received messages (direct and broadcast)
        received_messages = Message.query.join(
            MessageRead, Message.id == MessageRead.message_id
        ).filter(
            MessageRead.user_id == user_id
        ).all()

        # Sent direct messages should also appear in sender's chat list.
        sent_messages = Message.query.filter(
            Message.sender_id == user_id,
            Message.is_broadcast == False
        ).all()

        combined_by_id = {}
        for msg in received_messages:
            combined_by_id[msg.id] = msg
        for msg in sent_messages:
            combined_by_id[msg.id] = msg

        all_messages = sorted(combined_by_id.values(), key=lambda m: m.created_at, reverse=True)

        total = len(all_messages)
        total_pages = max(1, (total + per_page - 1) // per_page)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        page_items = all_messages[start_idx:end_idx]

        items = []
        for msg in page_items:
            sender = User.query.filter_by(user_id=msg.sender_id).first()
            recipient = User.query.filter_by(user_id=msg.recipient_id).first() if msg.recipient_id else None
            msg_read = MessageRead.query.filter_by(message_id=msg.id, user_id=user_id).first()

            # Sent messages should appear as read for the sender's own view.
            is_read = True if msg.sender_id == user_id else (msg_read.is_read if msg_read else False)

            items.append({
                "id": msg.id,
                "sender_id": msg.sender_id,
                "sender_name": sender.name if sender else "System",
                "recipient_id": msg.recipient_id,
                "recipient_name": recipient.name if recipient else None,
                "sender_role": sender.role if sender else None,
                "recipient_role": recipient.role if recipient else None,
                "title": msg.title,
                "content": msg.content,
                "type": msg.type,
                "message_type": msg.type,
                "is_read": is_read,
                "created_at": msg.created_at.isoformat() + 'Z',
                "is_broadcast": msg.is_broadcast
            })

        return jsonify({
            "success": True,
            "messages": items,
            "page": page,
            "total_pages": total_pages,
            "total": total
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/messages/<message_id>/read', methods=['POST'])
def mark_message_read(message_id):
    """Mark a message as read"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')

        msg_read = MessageRead.query.filter_by(message_id=message_id, user_id=user_id).first()
        if not msg_read:
            return jsonify({"error": "Message not found for user"}), 404

        msg_read.is_read = True
        msg_read.read_at = datetime.utcnow()
        db.session.commit()

        return jsonify({"success": True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/messages/<user_id>/with/<other_user_id>', methods=['GET'])
def get_conversation(user_id, other_user_id):
    """Get direct message conversation between two users"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = 50

        # Get all messages between these two users (either direction)
        messages = Message.query.filter(
            db.or_(
                db.and_(Message.sender_id == user_id, Message.recipient_id == other_user_id),
                db.and_(Message.sender_id == other_user_id, Message.recipient_id == user_id)
            ),
            Message.is_broadcast == False  # Only direct messages, not broadcasts
        ).order_by(Message.created_at.asc()).paginate(page=page, per_page=per_page)

        items = []
        for msg in messages.items:
            sender = User.query.filter_by(user_id=msg.sender_id).first()
            recipient = User.query.filter_by(user_id=msg.recipient_id).first()

            items.append({
                "id": msg.id,
                "sender_id": msg.sender_id,
                "sender_name": sender.name if sender else "Unknown",
                "sender_role": sender.role if sender else None,
                "recipient_id": msg.recipient_id,
                "recipient_name": recipient.name if recipient else "Unknown",
                "recipient_role": recipient.role if recipient else None,
                "content": msg.content,
                "type": msg.type,
                "message_type": msg.type,
                "created_at": msg.created_at.isoformat() + 'Z',
                "is_read": True  # Personal messages are always marked as read in UI
            })

        # Mark all messages from other user as read for current user
        for msg in messages.items:
            if msg.sender_id == other_user_id:
                msg_read = MessageRead.query.filter_by(message_id=msg.id, user_id=user_id).first()
                if msg_read:
                    msg_read.is_read = True
        db.session.commit()

        return jsonify({
            "success": True,
            "messages": items,
            "page": page,
            "total_pages": messages.pages,
            "total": messages.total
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/users/search', methods=['GET'])
def search_users():
    """Search for users by name or user_id"""
    try:
        query = request.args.get('q', '').strip()

        if not query or len(query) < 2:
            return jsonify({"success": True, "users": []}), 200

        # Search by name or user_id (case-insensitive)
        users = User.query.filter(
            db.or_(
                User.name.ilike(f'%{query}%'),
                User.user_id.ilike(f'%{query}%')
            ),
            User.is_active == True  # Only active users
        ).limit(10).all()

        users_list = [{
            "user_id": u.user_id,
            "name": u.name,
            "email": u.email,
            "role": u.role
        } for u in users]

        return jsonify({
            "success": True,
            "users": users_list
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --------- ALERT PREFERENCES ---------

@app.route('/api/alert-preferences/<user_id>', methods=['GET'])
def get_alert_preferences(user_id):
    """Get alert preferences for a user"""
    try:
        prefs = AlertPreference.query.filter_by(user_id=user_id).first()
        if not prefs:
            # Create default preferences if not exist
            prefs = AlertPreference(user_id=user_id)
            db.session.add(prefs)
            db.session.commit()

        return jsonify({
            "success": True,
            "preferences": {
                "alert_late_arrival": prefs.alert_late_arrival,
                "alert_approval_status": prefs.alert_approval_status,
                "alert_policy_violation": prefs.alert_policy_violation,
                "alert_announcements": prefs.alert_announcements,
                "alert_failed_scans": prefs.alert_failed_scans,
                "alert_suspicious_activity": prefs.alert_suspicious_activity,
                "enable_in_app_notifications": prefs.enable_in_app_notifications,
                "enable_push_notifications": prefs.enable_push_notifications,
                "quiet_hours_start": prefs.quiet_hours_start,
                "quiet_hours_end": prefs.quiet_hours_end
            }
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/alert-preferences/<user_id>', methods=['POST'])
def update_alert_preferences(user_id):
    """Update alert preferences for a user"""
    try:
        data = request.get_json()

        prefs = AlertPreference.query.filter_by(user_id=user_id).first()
        if not prefs:
            prefs = AlertPreference(user_id=user_id)
            db.session.add(prefs)

        # Update fields if provided
        if 'alert_late_arrival' in data:
            prefs.alert_late_arrival = data['alert_late_arrival']
        if 'alert_approval_status' in data:
            prefs.alert_approval_status = data['alert_approval_status']
        if 'alert_policy_violation' in data:
            prefs.alert_policy_violation = data['alert_policy_violation']
        if 'alert_announcements' in data:
            prefs.alert_announcements = data['alert_announcements']
        if 'alert_failed_scans' in data:
            prefs.alert_failed_scans = data['alert_failed_scans']
        if 'alert_suspicious_activity' in data:
            prefs.alert_suspicious_activity = data['alert_suspicious_activity']
        if 'enable_in_app_notifications' in data:
            prefs.enable_in_app_notifications = data['enable_in_app_notifications']
        if 'enable_push_notifications' in data:
            prefs.enable_push_notifications = data['enable_push_notifications']
        if 'quiet_hours_start' in data:
            prefs.quiet_hours_start = data['quiet_hours_start']
        if 'quiet_hours_end' in data:
            prefs.quiet_hours_end = data['quiet_hours_end']

        prefs.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({"success": True, "message": "Preferences updated"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# --------- PINNED ALERTS INBOX ---------

@app.route('/api/alerts/pinned/<user_id>', methods=['GET'])
def get_pinned_alerts(user_id):
    """Get all pinned/flagged alerts for a user"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = 20

        pinned = Notification.query.filter_by(
            user_id=user_id,
            is_pinned=True
        ).order_by(Notification.pinned_at.desc()).paginate(page=page, per_page=per_page)

        items = []
        for notif in pinned.items:
            items.append({
                "id": notif.id,
                "title": notif.title,
                "message": notif.message,
                "type": notif.type,
                "priority": notif.type if notif.type in ['info', 'warning', 'critical'] else 'info',
                "is_read": notif.is_read,
                "created_at": notif.created_at.isoformat() + 'Z',
                "pinned_at": (notif.pinned_at.isoformat() + 'Z') if notif.pinned_at else None
            })

        return jsonify({
            "success": True,
            "alerts": items,
            "page": page,
            "total_pages": pinned.pages,
            "total": pinned.total
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/alerts/<alert_id>/pin', methods=['POST'])
def pin_alert(alert_id):
    """Pin/flag an alert"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')

        alert = Notification.query.filter_by(id=alert_id, user_id=user_id).first()
        if not alert:
            return jsonify({"error": "Alert not found"}), 404

        alert.is_pinned = True
        alert.pinned_at = datetime.utcnow()
        db.session.commit()

        return jsonify({"success": True, "message": "Alert pinned"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/alerts/<alert_id>/unpin', methods=['POST'])
def unpin_alert(alert_id):
    """Unpin/unflag an alert"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')

        alert = Notification.query.filter_by(id=alert_id, user_id=user_id).first()
        if not alert:
            return jsonify({"error": "Alert not found"}), 404

        alert.is_pinned = False
        alert.pinned_at = None
        db.session.commit()

        return jsonify({"success": True, "message": "Alert unpinned"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/alerts/clear_all', methods=['POST'])
def clear_all_alerts_for_user():
    """Clear all pinned alerts for a user."""
    try:
        data = request.get_json() or {}
        user_id = data.get('user_id')

        if not user_id:
            return jsonify({"error": "Missing user_id"}), 400

        Notification.query.filter_by(user_id=user_id, is_pinned=True).update(
            {"is_pinned": False, "pinned_at": None},
            synchronize_session=False
        )
        db.session.commit()

        return jsonify({"success": True, "message": "All alerts cleared"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# --------- DARK MODE / THEME PREFERENCES ---------

# --------- ANTI-SPOOFING LIVENESS CHECK ---------

def generate_liveness_challenge():
    """Generate a random liveness challenge"""
    import random
    challenges = [
        {"type": "blink", "instructions": "Blink your eyes twice", "duration": 3},
        {"type": "head_turn", "instructions": "Turn your head left, then right", "duration": 5},
        {"type": "smile", "instructions": "Smile naturally", "duration": 3},
        {"type": "nod", "instructions": "Nod your head up and down", "duration": 4}
    ]
    return random.choice(challenges)

@app.route('/api/liveness/challenge/start', methods=['POST'])
def start_liveness_challenge():
    """Start a liveness challenge for face registration"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')

        if not user_id:
            return jsonify({"error": "Missing user_id"}), 400

        challenge = generate_liveness_challenge()

        # Create liveness challenge record
        liveness = LivenessChallenge(
            user_id=user_id,
            challenge_type=challenge['type'],
            challenge_instructions=challenge['instructions'],
            is_passed=None
        )
        db.session.add(liveness)
        db.session.commit()

        return jsonify({
            "success": True,
            "challenge_id": liveness.id,
            "challenge_type": challenge['type'],
            "instructions": challenge['instructions'],
            "duration_seconds": challenge['duration']
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/liveness/challenge/<challenge_id>/verify', methods=['POST'])
def verify_liveness_challenge(challenge_id):
    """Verify liveness challenge result"""
    try:
        data = request.get_json()
        is_passed = data.get('is_passed', False)

        challenge = LivenessChallenge.query.filter_by(id=challenge_id).first()
        if not challenge:
            return jsonify({"error": "Challenge not found"}), 404

        challenge.is_passed = is_passed
        challenge.completed_at = datetime.utcnow()

        if is_passed:
            challenge.attempt_count = 1
        else:
            challenge.attempt_count += 1

        db.session.commit()

        if is_passed:
            # Create notification for successful liveness check
            notif = Notification(
                user_id=challenge.user_id,
                title="✅ Liveness Check Passed",
                message="Your face registration passed the anti-spoofing verification",
                type="verification"
            )
            db.session.add(notif)
            db.session.commit()

        return jsonify({
            "success": True,
            "is_passed": is_passed,
            "message": "Liveness verification passed!" if is_passed else "Liveness verification failed. Try again.",
            "attempt_count": challenge.attempt_count
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# --------- DUPLICATE FACE DETECTION ---------

def check_duplicate_faces(new_encoding, threshold=0.60):
    """Check if new encoding matches existing faces (potential duplicate)"""
    try:
        all_users = User.query.filter(User.face_encoding != None).all()
        duplicates = []

        new_enc = np.array(new_encoding, dtype=np.float32).reshape(1, 128)

        for user in all_users:
            if user.face_encoding is None:
                continue

            try:
                # If it's a list of encodings, check all
                encodings_to_check = user.face_encoding if isinstance(user.face_encoding, list) else [user.face_encoding]

                for known_enc in encodings_to_check:
                    if known_enc is None:
                        continue

                    known = np.array(known_enc, dtype=np.float32).reshape(1, 128)
                    # Using OpenCV face recognizer (assumes face_system is initialized)
                    score = face_system.recognizer.match(known, new_enc, cv2.FaceRecognizerSF_FR_COSINE)

                    if score >= threshold:
                        duplicates.append({
                            "user_id": user.user_id,
                            "name": user.name,
                            "similarity_score": float(score)
                        })
            except:
                continue

        return duplicates
    except Exception as e:
        print(f"Duplicate face check error: {e}")
        return []

@app.route('/api/face/check-duplicates', methods=['POST'])
def check_duplicate_faces_endpoint():
    """Check if uploaded face is duplicate of any existing face"""
    try:
        if 'face_image' not in request.files:
            return jsonify({"error": "No face image provided"}), 400

        user_id = request.form.get('user_id')
        file = request.files['face_image']

        if not file or not user_id:
            return jsonify({"error": "Missing user_id or face_image"}), 400

        # Get face encoding
        encoding, _ = face_system.get_face_encoding(file.stream)

        if encoding is None:
            return jsonify({"error": "No face detected in image"}), 400

        # Check for duplicates
        duplicates = check_duplicate_faces(encoding, threshold=0.60)

        if duplicates:
            # Create duplicate alerts
            for dup in duplicates:
                alert = DuplicateFaceAlert(
                    primary_user_id=dup['user_id'],
                    duplicate_user_id=user_id,
                    similarity_score=dup['similarity_score'],
                    is_flagged=True
                )
                db.session.add(alert)

            db.session.commit()

            return jsonify({
                "success": False,
                "is_duplicate": True,
                "message": f"Face matches {len(duplicates)} existing registration(s)",
                "duplicates": duplicates
            }), 200

        return jsonify({
            "success": True,
            "is_duplicate": False,
            "message": "No duplicate faces detected"
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/duplicate-alerts/<admin_id>', methods=['GET'])
def get_duplicate_alerts(admin_id):
    """Get all duplicate face alerts for admin review"""
    try:
        # Verify admin
        admin = User.query.filter_by(user_id=admin_id, role='admin').first()
        if not admin:
            return jsonify({"error": "Unauthorized"}), 403

        page = request.args.get('page', 1, type=int)
        per_page = 20

        alerts = DuplicateFaceAlert.query.filter_by(is_flagged=True).order_by(
            DuplicateFaceAlert.created_at.desc()
        ).paginate(page=page, per_page=per_page)

        items = []
        for alert in alerts.items:
            primary = User.query.filter_by(user_id=alert.primary_user_id).first()
            duplicate = User.query.filter_by(user_id=alert.duplicate_user_id).first()

            items.append({
                "id": alert.id,
                "primary_user": {
                    "user_id": primary.user_id if primary else None,
                    "name": primary.name if primary else "Unknown"
                },
                "duplicate_user": {
                    "user_id": duplicate.user_id if duplicate else None,
                    "name": duplicate.name if duplicate else "Unknown"
                },
                "similarity_score": alert.similarity_score,
                "admin_notes": alert.admin_notes,
                "created_at": alert.created_at.isoformat(),
                "resolved_at": alert.resolved_at.isoformat() if alert.resolved_at else None
            })

        return jsonify({
            "success": True,
            "alerts": items,
            "page": page,
            "total_pages": alerts.pages,
            "total": alerts.total
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/duplicate-alerts/<alert_id>/resolve', methods=['POST'])
def resolve_duplicate_alert(alert_id):
    """Mark duplicate alert as resolved"""
    try:
        data = request.get_json()
        admin_id = data.get('admin_id')
        action = data.get('action')  # 'approve', 'reject', 'note'
        notes = data.get('notes', '')

        admin = User.query.filter_by(user_id=admin_id, role='admin').first()
        if not admin:
            return jsonify({"error": "Unauthorized"}), 403

        alert = DuplicateFaceAlert.query.filter_by(id=alert_id).first()
        if not alert:
            return jsonify({"error": "Alert not found"}), 404

        alert.is_flagged = False
        alert.resolved_at = datetime.utcnow()
        alert.admin_notes = notes

        # If action is reject, notify the duplicate user
        if action == 'reject':
            notif = Notification(
                user_id=alert.duplicate_user_id,
                title="⚠️ Duplicate Face Detected",
                message="Your face registration appears to match another user. Please contact admin.",
                type="warning"
            )
            db.session.add(notif)

        db.session.commit()

        return jsonify({"success": True, "message": "Alert resolved"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# --------- END NOTIFICATION API ---------

@app.route('/api/messages/<int:message_id>', methods=['DELETE'])
def delete_message(message_id):
    """Delete a specific message"""
    try:
        msg = Message.query.get(message_id)
        if not msg:
            return jsonify({"error": "Message not found"}), 404

        # Delete associated read receipts first to avoid foreign key constraints
        MessageRead.query.filter_by(message_id=message_id).delete()

        db.session.delete(msg)
        db.session.commit()

        return jsonify({"success": True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/admin/alerts/send', methods=['POST'])
def send_admin_notification_alert():
    """Route admin alerts directly to the Notification table (Alerts Tab)"""
    try:
        data = request.get_json()
        sender_id = data.get('sender_id')
        recipient_id = data.get('recipient_id')
        title = data.get('title')
        content = data.get('content')
        priority = str(data.get('priority', 'info')).lower()

        if not sender_id or not recipient_id or not title or not content:
            return jsonify({"error": "Missing required fields"}), 400

        if priority not in ['info', 'warning', 'critical']:
            priority = 'info'

        sender = User.query.filter_by(user_id=sender_id).first()
        if not sender or sender.role != 'admin':
            return jsonify({"error": "Unauthorized"}), 403

        recipient = User.query.filter_by(user_id=recipient_id).first()
        if not recipient:
            return jsonify({"error": "Recipient not found"}), 404

        # Create Notification directly so it goes to the Alerts tab
        notif = Notification(
            user_id=recipient_id,
            title=title,
            message=content,
            type=priority,
            is_pinned=True, # Auto-pin so it shows in the pinned alerts tab
            pinned_at=datetime.utcnow()
        )
        db.session.add(notif)
        db.session.commit()

        # Send FCM wakeup ping if token exists (Admin SDK)
        tokens = FCMToken.query.filter_by(user_id=recipient_id, is_active=True).all()
        for token in tokens:
            ping_device_via_fcm(token.fcm_token, "wakeup")

        return jsonify({"success": True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
        
# Chat Autocomplete Endpoint
@app.route('/api/users/search', methods=['GET'])
def search_all_users():
    """Get all active users for chat autocomplete"""
    try:
        users = User.query.all()
        results = [
            {
                "user_id": u.user_id,
                "name": u.name,
                "role": u.role
            }
            for u in users if u.is_active is not False # Allows True and None (existing users)
        ]
        return jsonify({"success": True, "users": results}), 200
    except Exception as e:
        print(f"User Search API Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

if __name__ == '__main__':
    # Reverting to standard HTTP (No SSL)
    # Note: Mobile browsers may block camera access on http://192.168.x.x
    app.run(debug=True, host='0.0.0.0', port=5000)
